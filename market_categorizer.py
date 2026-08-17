#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
market_categorizer.py -- Market Categorization Engine
=====================================================

Purpose
-------
Classify every market in the bank's Aug-2026 branch footprint (including the
branches added by the January-2026 Synovus merger) by SIZE TIER and GROWTH TIER,
using only locally cached public data. THIS SCRIPT MAKES NO NETWORK CALLS.

Unit of analysis
----------------
  * one row per footprint CBSA (Metropolitan or Micropolitan Statistical Area),
  * one row per footprint county that is not part of any CBSA
    (market_type = 'Rural county').

Size tier (Census Vintage-2025 population, 2023 OMB delineations)
-----------------------------------------------------------------
  Major Metro     : Metropolitan Statistical Area, pop2025 >= 1,000,000
  Secondary Metro : Metropolitan Statistical Area, 250,000 <= pop2025 <= 999,999
  Small Metro     : Metropolitan Statistical Area, pop2025 < 250,000
  Micropolitan    : Micropolitan Statistical Area (any size)
  Rural           : non-CBSA county

Growth indicators (5)
---------------------
  1. pop_cagr     2020 -> 2025 population CAGR                        (Census PEP)
  2. netmig_rate  cumulative net DOMESTIC migration 2021-2025 / pop2020 (Census PEP)
  3. gdp_cagr     2019 -> 2024 real GDP CAGR                          (BEA CAGDP9)
  4. emp_cagr     2019 -> 2024 QCEW annual-average employment CAGR    (BLS QCEW)
  5. pcpi_cagr    2019 -> 2024 per-capita personal income CAGR        (BEA CAINC1)

Standardization
---------------
  * CBSA markets: each of pop/netmig/gdp/pcpi is z-scored against the distribution
    of ALL 925 US metro+micro CBSAs (unweighted; every CBSA counts once).
    The mean and standard deviation are computed AFTER winsorizing that national
    distribution at the 1st and 99th percentiles, so a handful of extreme small
    areas cannot inflate the spread. The market's own (unwinsorized) value is
    then scored against those robust moments.
  * Rural county markets are z-scored the same way against the distribution of
    ALL US counties (Census/BEA county universes).
  * emp_cagr has no national CBSA distribution (QCEW was pulled only for
    footprint areas). Its z-score is therefore
        z_emp = (area_emp_cagr - US_emp_cagr) / std(emp_cagr over footprint areas)
    where the standard deviation uses the winsorized footprint distribution.
    This is a WEAKER benchmark than the other four -- see methodology.md.

Composite and growth tier
-------------------------
  composite_z = 0.225*z_pop + 0.225*z_netmig + 0.225*z_gdp + 0.225*z_pcpi
              + 0.100*z_emp
  Weights are renormalized over whichever indicators are present; any renorm is
  recorded in data_flags.
  High   : composite_z >= +0.33
  Low    : composite_z <= -0.33
  Medium : otherwise

Outputs (written to out/)
-------------------------
  market_scores.csv, market_scores.xlsx, methodology.md
"""

from __future__ import annotations

import json
import math
import os
from datetime import date

import numpy as np
import pandas as pd

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")
OUT = os.path.join(BASE, "out")
os.makedirs(OUT, exist_ok=True)

WEIGHTS = {           # composite weights, renormalized when an indicator is missing
    "pop_cagr": 0.225,
    "netmig_rate": 0.225,
    "gdp_cagr": 0.225,
    "pcpi_cagr": 0.225,
    "emp_cagr": 0.100,
}
HIGH_CUT, LOW_CUT = 0.33, -0.33
POP_YEARS = 5         # 2020 -> 2025
ECON_YEARS = 5        # 2019 -> 2024
WINSOR = (1.0, 99.0)  # percentiles used to trim the benchmark distribution

FOOTPRINT_STATES = {"TN", "VA", "DC", "KY", "NC", "SC", "GA", "FL", "AL"}


def p(*parts: str) -> str:
    return os.path.join(DATA, *parts)


def cagr(v0, v1, years: int):
    """Compound annual growth rate; None if either endpoint is missing/non-positive."""
    if v0 is None or v1 is None:
        return None
    try:
        v0, v1 = float(v0), float(v1)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(v0) or not np.isfinite(v1) or v0 <= 0 or v1 <= 0:
        return None
    return (v1 / v0) ** (1.0 / years) - 1.0


def winsorized_moments(series: pd.Series):
    """(mean, std) of a distribution after clipping at the 1st/99th percentiles."""
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) < 3:
        return (float("nan"), float("nan"), 0)
    lo, hi = np.percentile(s, WINSOR[0]), np.percentile(s, WINSOR[1])
    w = s.clip(lo, hi)
    # ddof=0: these are the full observed populations of CBSAs / counties, not samples
    return (float(w.mean()), float(w.std(ddof=0)), int(len(s)))


# --------------------------------------------------------------------------- #
# 1. Load the footprint (markets we must score)
# --------------------------------------------------------------------------- #
def load_footprint():
    """
    footprint_msa.csv is the branch-level footprint aggregated to market level.
    Columns: cbsa_code, cbsa_title, metro_micro, county_fips, branches,
             branches_acquired, deposits_2025_kusd, deposits_2020_kusd,
             states, in_footprint_states
    Rural rows carry cbsa_code='RURAL' and the county FIPS in county_fips.
    """
    fm = pd.read_csv(
        p("footprint_msa.csv"),
        dtype={"cbsa_code": str, "county_fips": str},
    )
    fm["county_fips"] = fm["county_fips"].fillna("").str.zfill(5).replace("00000", "")
    fm["is_rural"] = fm["cbsa_code"].eq("RURAL")
    fm["key"] = np.where(fm["is_rural"], fm["county_fips"], fm["cbsa_code"])
    return fm


def load_office_types():
    """
    FDIC office service type per UNINUM, from the same cached August-2026 office
    file the footprint was built from. Only FULL SERVICE offices take deposits and
    file a Summary of Deposits record; LIMITED SERVICE offices (loan production,
    messenger, drive-thru, administrative, mobile/seasonal, other) do not.
    """
    src = pd.read_csv(
        p("uploaded_branch_locations.csv"), encoding="utf-8-sig", dtype=str
    )
    src["uninum"] = src["UNINUM"].str.strip()
    src["servtype_desc"] = src["SERVTYPE_DESC"].fillna("").str.strip().str.upper()
    src["full_service"] = src["servtype_desc"].str.startswith("FULL SERVICE")
    src["established"] = pd.to_datetime(src["ESTYMD"], errors="coerce")
    return src[["uninum", "servtype_desc", "full_service", "established"]]


def load_branches():
    br = pd.read_csv(
        p("footprint.csv"),
        dtype={"county_fips": str, "cbsa_code": str, "uninum": str},
    )
    br["county_fips"] = br["county_fips"].str.zfill(5)
    for c in ("deposits_2025_kusd", "deposits_2020_kusd"):
        br[c] = pd.to_numeric(br[c], errors="coerce")
    br["acquired_2026"] = br["acquired_2026"].astype(str).str.lower().eq("true")
    br["key"] = np.where(br["cbsa_code"].eq("RURAL"), br["county_fips"], br["cbsa_code"])
    br["uninum"] = br["uninum"].astype(str).str.strip()
    br = br.merge(load_office_types(), on="uninum", how="left")
    br["full_service"] = br["full_service"].fillna(False)
    br["servtype_desc"] = br["servtype_desc"].fillna("UNKNOWN")
    br["no_sod2025_match"] = br["flags"].fillna("").str.contains("no_sod2025_match")
    return br


# --------------------------------------------------------------------------- #
# 2. Build the national CBSA benchmark table (all 925 metro + micro CBSAs)
# --------------------------------------------------------------------------- #
def build_cbsa_universe():
    pop = pd.read_csv(p("pop_cbsa.csv"), dtype={"cbsa_code": str})
    pop["is_metro"] = pop["lsad"].str.startswith("Metropolitan")

    u = pop[["cbsa_code", "cbsa_title", "lsad", "is_metro",
             "popestimate2020", "popestimate2025",
             "domesticmig_2021_2025"]].copy()
    u["pop_cagr"] = [cagr(a, b, POP_YEARS)
                     for a, b in zip(u.popestimate2020, u.popestimate2025)]
    u["netmig_rate"] = u["domesticmig_2021_2025"] / u["popestimate2020"]

    # --- real GDP (BEA CAGDP9, chained 2017$; MSA/micro built by county aggregation)
    gdp = pd.read_csv(p("gdp_msa.csv"), dtype={"geofips": str})
    gdp = gdp[gdp.year.isin([2019, 2024])].pivot_table(
        index="geofips", columns="year", values="real_gdp_k2017usd", aggfunc="first")
    gdp["gdp_cagr"] = [cagr(a, b, ECON_YEARS) for a, b in zip(gdp.get(2019), gdp.get(2024))]
    u = u.merge(gdp[["gdp_cagr"]], left_on="cbsa_code", right_index=True, how="left")

    # --- per-capita personal income (BEA CAINC1, nominal $)
    pcpi = pd.read_csv(p("pcpi_msa.csv"), dtype={"geofips": str})
    pcpi = pcpi[pcpi.year.isin([2019, 2024])].pivot_table(
        index="geofips", columns="year",
        values="per_capita_personal_income_usd", aggfunc="first")
    pcpi["pcpi_cagr"] = [cagr(a, b, ECON_YEARS) for a, b in zip(pcpi.get(2019), pcpi.get(2024))]
    u = u.merge(pcpi[["pcpi_cagr"]], left_on="cbsa_code", right_index=True, how="left")

    return u.set_index("cbsa_code")


def build_county_universe():
    """National county benchmark distribution used for Rural-county markets."""
    pop = pd.read_csv(p("pop_county.csv"), dtype={"fips": str})
    pop["fips"] = pop["fips"].str.zfill(5)
    u = pop[["fips", "county_name", "state_name", "popestimate2020",
             "popestimate2025", "domesticmig_2021_2025"]].copy()
    u["pop_cagr"] = [cagr(a, b, POP_YEARS)
                     for a, b in zip(u.popestimate2020, u.popestimate2025)]
    u["netmig_rate"] = u["domesticmig_2021_2025"] / u["popestimate2020"]

    gdp = pd.read_csv(p("gdp_county.csv"), dtype={"geofips": str})
    gdp["geofips"] = gdp["geofips"].str.zfill(5)
    gdp = gdp[gdp.year.isin([2019, 2024])].pivot_table(
        index="geofips", columns="year", values="real_gdp_k2017usd", aggfunc="first")
    gdp["gdp_cagr"] = [cagr(a, b, ECON_YEARS) for a, b in zip(gdp.get(2019), gdp.get(2024))]
    u = u.merge(gdp[["gdp_cagr"]], left_on="fips", right_index=True, how="left")

    pcpi = pd.read_csv(p("pcpi_county.csv"), dtype={"geofips": str})
    pcpi["geofips"] = pcpi["geofips"].str.zfill(5)
    pcpi = pcpi[pcpi.year.isin([2019, 2024])].pivot_table(
        index="geofips", columns="year",
        values="per_capita_personal_income_usd", aggfunc="first")
    pcpi["pcpi_cagr"] = [cagr(a, b, ECON_YEARS) for a, b in zip(pcpi.get(2019), pcpi.get(2024))]
    u = u.merge(pcpi[["pcpi_cagr"]], left_on="fips", right_index=True, how="left")

    return u.set_index("fips")


# --------------------------------------------------------------------------- #
# 3. QCEW employment (footprint areas only)
# --------------------------------------------------------------------------- #
def build_employment():
    em = pd.read_csv(p("emp_msa.csv"), dtype={"cbsa_code": str})
    em = em.pivot_table(index="cbsa_code", columns="year",
                        values="annual_avg_emplvl", aggfunc="first")
    ec = pd.read_csv(p("emp_county.csv"), dtype={"county_fips": str})
    ec["county_fips"] = ec["county_fips"].str.zfill(5)
    ec = ec.pivot_table(index="county_fips", columns="year",
                        values="annual_avg_emplvl", aggfunc="first")
    both = pd.concat([em, ec])
    emp = {k: cagr(r.get(2019), r.get(2024), ECON_YEARS) for k, r in both.iterrows()}

    us = pd.read_csv(p("emp_us.csv"))
    us = us.set_index("year")["annual_avg_emplvl"]
    us_cagr = cagr(us.loc[2019], us.loc[2024], ECON_YEARS)
    return emp, us_cagr


# --------------------------------------------------------------------------- #
# 4. National reference CAGRs (for the benchmarks sheet / documentation)
# --------------------------------------------------------------------------- #
def national_reference(us_emp_cagr):
    pop = pd.read_csv(p("pop_us.csv")).iloc[0]
    gdp = pd.read_csv(p("gdp_us.csv")).set_index("year")["real_gdp_k2017usd"]
    pcpi = pd.read_csv(p("pcpi_us.csv")).set_index("year")["per_capita_personal_income_usd"]
    return {
        # NOTE: national DOMESTIC migration is 0 by construction; the meaningful
        # national migration figure is NET migration (incl. international).
        "pop_cagr": cagr(pop.popestimate2020, pop.popestimate2025, POP_YEARS),
        "netmig_rate": float(pop.netmig_2021_2025) / float(pop.popestimate2020),
        "gdp_cagr": cagr(gdp.loc[2019], gdp.loc[2024], ECON_YEARS),
        "pcpi_cagr": cagr(pcpi.loc[2019], pcpi.loc[2024], ECON_YEARS),
        "emp_cagr": us_emp_cagr,
    }


# --------------------------------------------------------------------------- #
# 5. Size tier
# --------------------------------------------------------------------------- #
def size_tier(is_rural: bool, is_metro: bool, pop2025) -> str:
    if is_rural:
        return "Rural"
    if not is_metro:
        return "Micropolitan"
    if pop2025 is None or not np.isfinite(pop2025):
        return "Small Metro"
    if pop2025 >= 1_000_000:
        return "Major Metro"
    if pop2025 >= 250_000:
        return "Secondary Metro"
    return "Small Metro"


# --------------------------------------------------------------------------- #
# 6. Main build
# --------------------------------------------------------------------------- #
def main() -> dict:
    fm = load_footprint()
    br = load_branches()
    cbsa_u = build_cbsa_universe()
    cnty_u = build_county_universe()
    emp_map, us_emp_cagr = build_employment()
    natl = national_reference(us_emp_cagr)

    # ---- benchmark moments -------------------------------------------------
    IND4 = ["pop_cagr", "netmig_rate", "gdp_cagr", "pcpi_cagr"]
    cbsa_mom = {i: winsorized_moments(cbsa_u[i]) for i in IND4}
    cnty_mom = {i: winsorized_moments(cnty_u[i]) for i in IND4}

    # emp: footprint-only spread, centered on the US CAGR
    fp_emp = pd.Series([emp_map.get(k) for k in fm["key"]], dtype="float64").dropna()
    _, emp_sd, emp_n = winsorized_moments(fp_emp)
    emp_mom = (us_emp_cagr, emp_sd, emp_n)

    # ---- deposit context per market ---------------------------------------
    dep_ctx = {}
    for key, grp in br.groupby("key"):
        both = grp.dropna(subset=["deposits_2025_kusd", "deposits_2020_kusd"])
        # ---- Endpoint guard, now SYMMETRIC on both sides of a reallocation ----
        #
        # A branch that reports deposits in 2020 and a literal $0 in 2025 has had
        # its book moved to another office of record; it did not lose its
        # deposits. Dropping it from the 2020 base is correct ONLY if the office
        # that received the book is dropped from the 2025 numerator too.
        #
        # 2026-08-17 fix. The previous version dropped only the SOURCE. In
        # Nashville that removed UNINUM 82012's $8.27bn from the 2020 base while
        # keeping NASHVILLE YARDS (UNINUM 465213, $42,947k -> $13,659,117k) --
        # the receiving office -- in the 2025 numerator, and printed a
        # "same-branch" CAGR of 32.4% against an underlying ~5-8%. Albany GA had
        # the same one-sided treatment via the Five Points branch.
        #
        # The fix pairs the two sides instead of dropping either: when an office
        # in the SAME market absorbed at least the whole vanished book (absolute
        # increase >= the reallocated-out total), the source's 2020 balance is
        # netted back into the base, so source and recipient are compared as one
        # continuing unit. Nothing is thrown away and the ratio is symmetric.
        # Nashville 32.4% -> 8.5%; Albany GA 3.8% -> 1.6%.
        reallocated = both[(both["deposits_2020_kusd"] > 0)
                           & (both["deposits_2025_kusd"] <= 0)]
        both = both[(both["deposits_2020_kusd"] > 0) & (both["deposits_2025_kusd"] > 0)]
        realloc_out = float(reallocated["deposits_2020_kusd"].sum())
        recipients = both[(both["deposits_2025_kusd"] - both["deposits_2020_kusd"])
                          >= realloc_out] if realloc_out > 0 else both.iloc[0:0]
        d25, d20 = both["deposits_2025_kusd"].sum(), both["deposits_2020_kusd"].sum()
        paired = realloc_out if len(recipients) else 0.0
        d20 += paired
        n_new = int(grp["deposits_2020_kusd"].isna().sum())
        if len(both) and d20 > 0:
            g = (d25 / d20) ** (1 / 5) - 1
            txt = (f"same-branch 2020->2025 deposit CAGR {g*100:.1f}% "
                   f"({len(both)} of {len(grp)} branches matched both vintages "
                   f"with deposits at both endpoints)")
        else:
            g = None
            txt = f"no branch matched in both 2020 and 2025 SOD ({len(grp)} branches)"
        if n_new:
            txt += f"; {n_new} branch(es) with no 2020 SOD deposits"
        if len(reallocated):
            excl = int(realloc_out)
            if paired:
                rn = ", ".join(f"{r.offname} (uninum {r.uninum}, "
                               f"{int(r.deposits_2020_kusd):,}k->"
                               f"{int(r.deposits_2025_kusd):,}k)"
                               for r in recipients.itertuples())
                txt += (f"; {len(reallocated)} branch(es) reported $0 in 2025 after an "
                        f"internal reallocation ({excl:,}k of 2020 deposits) and are "
                        f"PAIRED WITH the receiving office(s) in the same market -- "
                        f"{rn} -- so both sides of the move sit on the same side of "
                        f"the ratio")
            else:
                txt += (f"; excludes {len(reallocated)} branch(es) reporting $0 in 2025 "
                        f"after an internal reallocation ({excl:,}k of 2020 deposits); "
                        f"NO office in this market absorbed a matching amount, so the "
                        f"book appears to have left the market -- read this CAGR as a "
                        f"surviving-branch figure only")
        dep_ctx[key] = (txt, g)

    # ---- office mix per market (only FULL SERVICE offices take deposits) ------
    fs_by_key = br.groupby("key")["full_service"].sum().astype(int).to_dict()

    # ---- score each market -------------------------------------------------
    rows = []
    for _, m in fm.iterrows():
        key, rural = m["key"], bool(m["is_rural"])
        flags = []

        if rural:
            if key in cnty_u.index:
                src = cnty_u.loc[key]
                name = f"{src['county_name']}, {src['state_name']}"
            else:
                src, name = None, m["cbsa_title"]
                flags.append("county_not_in_national_universe")
            mom, bench = cnty_mom, "US counties"
            market_type, is_metro = "Rural county", False
        else:
            if key in cbsa_u.index:
                src = cbsa_u.loc[key]
                name = src["cbsa_title"]
            else:
                src, name = None, m["cbsa_title"]
                flags.append("cbsa_not_in_national_universe")
            mom, bench = cbsa_mom, "US metro+micro CBSAs"
            is_metro = bool(src["is_metro"]) if src is not None else \
                m["metro_micro"].startswith("Metro")
            market_type = "Metropolitan Statistical Area" if is_metro \
                else "Micropolitan Statistical Area"

        pop2025 = float(src["popestimate2025"]) if src is not None else float("nan")
        tier_size = size_tier(rural, is_metro, pop2025)

        vals, zs = {}, {}
        for ind in IND4:
            v = float(src[ind]) if src is not None and pd.notna(src[ind]) else None
            vals[ind] = v
            mu, sd, _ = mom[ind]
            zs[ind] = (v - mu) / sd if (v is not None and sd and np.isfinite(sd)) else None
            if v is None:
                flags.append(f"missing:{ind}")

        e = emp_map.get(key)
        vals["emp_cagr"] = e
        zs["emp_cagr"] = ((e - emp_mom[0]) / emp_mom[1]) if (e is not None and emp_mom[1]) else None
        if e is None:
            flags.append("missing:emp_cagr")

        present = [i for i in WEIGHTS if zs[i] is not None]
        wsum = sum(WEIGHTS[i] for i in present)
        composite = sum(WEIGHTS[i] * zs[i] for i in present) / wsum if wsum else None
        if len(present) < len(WEIGHTS) and present:
            flags.append(f"weights_renormalized_over:{'+'.join(present)}")

        tier_growth = ("High" if composite >= HIGH_CUT
                       else "Low" if composite <= LOW_CUT
                       else "Medium") if composite is not None else "Unscored"

        ctx, _g = dep_ctx.get(key, ("no branch deposit data", None))
        n_offices = int(m["branches"])
        n_full = int(fs_by_key.get(key, 0))
        if n_full == 0:
            flags.append("no_full_service_office:limited_service_presence_only")
        elif n_full < n_offices:
            flags.append(f"limited_service_offices:{n_offices - n_full}")
        if m["branches_acquired"] == m["branches"]:
            flags.append("all_branches_from_2026_synovus_merger")
        elif m["branches_acquired"] > 0:
            flags.append("mixed_legacy_and_merger_branches")

        rows.append({
            "cbsa_code_or_fips": key,
            "market_name": name,
            "states": m["states"],
            "market_type": market_type,
            "size_tier": tier_size,
            "growth_tier": tier_growth,
            "composite_z": round(composite, 4) if composite is not None else None,
            "pop_cagr": vals["pop_cagr"], "z_pop_cagr": zs["pop_cagr"],
            "netmig_rate": vals["netmig_rate"], "z_netmig_rate": zs["netmig_rate"],
            "gdp_cagr": vals["gdp_cagr"], "z_gdp_cagr": zs["gdp_cagr"],
            "emp_cagr": vals["emp_cagr"], "z_emp_cagr": zs["emp_cagr"],
            "pcpi_cagr": vals["pcpi_cagr"], "z_pcpi_cagr": zs["pcpi_cagr"],
            "pop_2025": int(pop2025) if np.isfinite(pop2025) else None,
            "pinnacle_offices": n_offices,
            "pinnacle_full_service_branches": n_full,
            "pinnacle_limited_service_offices": n_offices - n_full,
            "branches_acquired_2026": int(m["branches_acquired"]),
            "office_file_deposits_2025_kusd": int(m["deposits_2025_kusd"]),
            "deposit_growth_context": ctx,
            "in_footprint_states": bool(m["in_footprint_states"]),
            "benchmark_universe": bench,
            "data_flags": "|".join(flags),
        })

    df = pd.DataFrame(rows)
    for c in ["pop_cagr", "netmig_rate", "gdp_cagr", "emp_cagr", "pcpi_cagr",
              "z_pop_cagr", "z_netmig_rate", "z_gdp_cagr", "z_emp_cagr", "z_pcpi_cagr"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").round(6)
    df = df.sort_values(["office_file_deposits_2025_kusd", "pinnacle_full_service_branches",
                         "pinnacle_offices"],
                        ascending=[False, False, False]).reset_index(drop=True)

    # ---- MODULE 10: franchise-stage dimension (appended; see section 8) -----
    df, realloc = attach_franchise(df, br)
    df = df[[c for c in df.columns if c != "data_flags"] + ["data_flags"]]

    csv_path = os.path.join(OUT, "market_scores.csv")
    df.to_csv(csv_path, index=False)

    bench_df = benchmarks_frame(cbsa_mom, cnty_mom, emp_mom, natl, cbsa_u, cnty_u)
    write_methodology(df, natl, cbsa_mom, cnty_mom, emp_mom,
                      office_diagnostics(br, cnty_u), realloc)
    write_xlsx(df, bench_df)

    return {"df": df, "natl": natl, "bench": bench_df, "realloc": realloc,
            "cbsa_mom": cbsa_mom, "cnty_mom": cnty_mom, "emp_mom": emp_mom}


# --------------------------------------------------------------------------- #
# 7. National benchmark table
# --------------------------------------------------------------------------- #
def benchmarks_frame(cbsa_mom, cnty_mom, emp_mom, natl, cbsa_u, cnty_u):
    recs = []
    for ind in ["pop_cagr", "netmig_rate", "gdp_cagr", "pcpi_cagr"]:
        for label, mom, uni in (("US metro+micro CBSAs (n=925)", cbsa_mom, cbsa_u),
                                ("US counties", cnty_mom, cnty_u)):
            mu, sd, n = mom[ind]
            s = pd.to_numeric(uni[ind], errors="coerce").dropna()
            recs.append({
                "indicator": ind, "benchmark_universe": label, "n_areas": n,
                "us_value": natl[ind],
                "winsorized_mean": mu, "winsorized_std": sd,
                "p10": float(np.percentile(s, 10)), "median": float(s.median()),
                "p90": float(np.percentile(s, 90)),
                "min": float(s.min()), "max": float(s.max()),
            })
    mu, sd, n = emp_mom
    recs.append({
        "indicator": "emp_cagr",
        "benchmark_universe": "footprint areas only (QCEW not pulled nationally)",
        "n_areas": n, "us_value": natl["emp_cagr"],
        "winsorized_mean": mu, "winsorized_std": sd,
        "p10": None, "median": None, "p90": None, "min": None, "max": None,
    })
    return pd.DataFrame(recs)


# --------------------------------------------------------------------------- #
# 8. Excel writer
# --------------------------------------------------------------------------- #
def write_xlsx(df: pd.DataFrame, bench: pd.DataFrame):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)

    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"

    for rec in df.itertuples(index=False):
        ws.append([None if (isinstance(v, float) and math.isnan(v)) else v for v in rec])

    pct_cols = {"pop_cagr", "netmig_rate", "gdp_cagr", "emp_cagr", "pcpi_cagr",
                "deposit_share_2025", "deposit_share_2020", "branch_share_2025",
                "branch_change_pct"}
    num2_cols = ({c for c in df.columns if c.startswith("z_")}
                 | {"composite_z", "branch_effectiveness_index", "share_delta_5yr_bps"})
    int_cols = {"pop_2025", "pinnacle_offices", "pinnacle_full_service_branches",
                "pinnacle_limited_service_offices", "branches_acquired_2026",
                "office_file_deposits_2025_kusd", "branches_2020", "branches_2025_sod",
                "net_branch_change_5yr", "proforma_deposits_2025_kusd",
                "proforma_deposits_2020_kusd", "market_deposits_2025_kusd",
                "market_deposits_2020_kusd", "market_branches_2025", "n_institutions",
                "book_moved_out_kusd", "book_moved_in_kusd"}
    widths = {"market_name": 42, "states": 10, "market_type": 30, "size_tier": 16,
              "growth_tier": 12, "deposit_growth_context": 62, "data_flags": 46,
              "benchmark_universe": 24, "cbsa_code_or_fips": 16,
              "office_file_deposits_2025_kusd": 18, "franchise_stage": 30,
              "stage_detail": 88, "market_deposits_2025_kusd": 20,
              "proforma_deposits_2025_kusd": 20, "proforma_deposits_2020_kusd": 20,
              "market_deposits_2020_kusd": 20}

    tier_fill = {"High": PatternFill("solid", fgColor="C6EFCE"),
                 "Medium": PatternFill("solid", fgColor="FFEB9C"),
                 "Low": PatternFill("solid", fgColor="FFC7CE")}
    gcol = list(df.columns).index("growth_tier") + 1

    for i, col in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(col, 13)
        fmt = ("0.00%" if col in pct_cols else
               "0.00" if col in num2_cols else
               "#,##0" if col in int_cols else None)
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = fmt
    for r in range(2, ws.max_row + 1):
        f = tier_fill.get(ws.cell(row=r, column=gcol).value)
        if f:
            ws.cell(row=r, column=gcol).fill = f
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"

    # ---- National_Benchmarks ----
    wb2 = wb.create_sheet("National_Benchmarks")
    wb2.append(list(bench.columns))
    for c in range(1, len(bench.columns) + 1):
        wb2.cell(row=1, column=c).fill = hdr_fill
        wb2.cell(row=1, column=c).font = hdr_font
    for rec in bench.itertuples(index=False):
        wb2.append([None if (isinstance(v, float) and math.isnan(v)) else v for v in rec])
    for i, col in enumerate(bench.columns, start=1):
        wb2.column_dimensions[get_column_letter(i)].width = \
            40 if col == "benchmark_universe" else 16
        if col not in ("indicator", "benchmark_universe", "n_areas"):
            for r in range(2, wb2.max_row + 1):
                wb2.cell(row=r, column=i).number_format = "0.000%"

    # ---- Franchise_Stage (MODULE 10) ----
    if "franchise_stage" in df.columns:
        write_franchise_sheet(wb, df)

    # ---- Methodology ----
    wm = wb.create_sheet("Methodology")
    wm.column_dimensions["A"].width = 128
    with open(os.path.join(OUT, "methodology.md"), encoding="utf-8") as fh:
        for line in fh.read().splitlines():
            wm.append([line])
    for r in range(1, wm.max_row + 1):
        cell = wm.cell(row=r, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if str(cell.value).startswith("#"):
            cell.font = Font(bold=True)

    wb.save(os.path.join(OUT, "market_scores.xlsx"))


# --------------------------------------------------------------------------- #
# 9. Methodology write-up
# --------------------------------------------------------------------------- #
def office_diagnostics(br, cnty_u):
    """
    Facts about the office file and the county benchmark universe that the
    methodology write-up states explicitly. Computed, never hard-coded, so the
    prose cannot drift away from the data the run actually used.
    """
    mix = br["servtype_desc"].value_counts().to_dict()
    nom = br[br["no_sod2025_match"]]
    est = pd.to_datetime(nom["established"], errors="coerce")
    cnty_nan = {i: int(cnty_u[i].isna().sum())
                for i in ("pop_cagr", "netmig_rate", "gdp_cagr", "pcpi_cagr")}
    return {
        "n_offices": int(len(br)),
        "n_full": int(br["full_service"].sum()),
        "n_limited": int((~br["full_service"]).sum()),
        "mix": mix,
        "limited_mix": {k: v for k, v in sorted(mix.items(), key=lambda kv: -kv[1])
                        if not k.startswith("FULL SERVICE")},
        "n_no_sod": int(len(nom)),
        "no_sod_limited": int((~nom["full_service"]).sum()),
        "no_sod_full": int(nom["full_service"].sum()),
        "no_sod_new": int((est > pd.Timestamp("2025-06-30")).sum()),
        "no_sod_pre2020": int((est < pd.Timestamp("2020-01-01")).sum()),
        "county_rows": int(len(cnty_u)),
        "county_nan": cnty_nan,
        "realloc": int(((br["deposits_2020_kusd"] > 0)
                        & (br["deposits_2025_kusd"] <= 0)).sum()),
    }


def write_methodology(df, natl, cbsa_mom, cnty_mom, emp_mom, od, realloc=None):
    def pc(x):
        return "n/a" if x is None or not np.isfinite(x) else f"{x*100:.2f}%"

    sizes = df["size_tier"].value_counts().to_dict()
    tiers = df["growth_tier"].value_counts().to_dict()

    zero_fs = df[df["pinnacle_full_service_branches"] == 0]
    zero_fs_txt = "; ".join(
        f"{r.cbsa_code_or_fips} {r.market_name} ({r.pinnacle_offices} office(s), "
        f"{r.size_tier} / {r.growth_tier})" for r in zero_fs.itertuples()
    ) or "none"
    limited_txt = ", ".join(f"{v} {k.replace('LIMITED SERVICE - ', '').lower()}"
                            for k, v in od["limited_mix"].items())
    cnty_n = {i: cnty_mom[i][2] for i in cnty_mom}

    # Office-file vs pro-forma deposit-column reconciliation (section 7).
    _of_tot = _pf_tot = _of_gap = _of_n = 0
    _of_list = "none"
    if "proforma_deposits_2025_kusd" in df.columns:
        _of = pd.to_numeric(df["office_file_deposits_2025_kusd"], errors="coerce").fillna(0)
        _pf = pd.to_numeric(df["proforma_deposits_2025_kusd"], errors="coerce").fillna(0)
        _of_tot, _pf_tot = int(_of.sum()), int(_pf.sum())
        _of_gap = _pf_tot - _of_tot
        _d = df.assign(gap_kusd=(_of - _pf).astype(int))
        _d = _d[(_d["gap_kusd"] != 0) & (_d["cbsa_code_or_fips"].astype(str) != "14540")]
        _of_n = int((( _of - _pf) != 0).sum())
        _of_list = "; ".join(f"{r.market_name} {r.gap_kusd:+,}k"
                             for r in _d.sort_values("gap_kusd").itertuples()) or "none"

    txt = f"""# Market Categorization Methodology

Generated {date.today().isoformat()} by `market_categorizer.py`. The scoring script
performs NO network calls; every input is a cached extract in `data/`.

## 1. Unit of analysis

One row per market:

* **CBSA markets** - each Metropolitan or Micropolitan Statistical Area (2023 OMB
  delineations, Sept-2023 vintage) in which the bank operates at least one office
  (see "What counts as a branch" below).
* **Rural county markets** - each footprint county that belongs to no CBSA;
  `market_type = 'Rural county'`.

The footprint is the bank's authoritative August-2026 FDIC office list
({od['n_offices']} offices), which already includes the offices added in the
January-2026 Synovus merger.

**What counts as a "branch".** The office file is a list of *FDIC offices*, not a
list of deposit-taking branches. By `SERVTYPE_DESC`, {od['n_full']} of the
{od['n_offices']} offices are FULL SERVICE (brick-and-mortar or retail) and
therefore take deposits and file a Summary of Deposits record; the other
{od['n_limited']} are LIMITED SERVICE ({limited_txt}) and take no deposits. Every
market row reports all three counts - `pinnacle_offices`,
`pinnacle_full_service_branches`, `pinnacle_limited_service_offices` - and the
word "branch" in this document means a **full-service, deposit-taking** office.
A market is included if the bank operates at least one office of any type in it,
so presence is not the same thing as a deposit franchise: {len(zero_fs)} market(s)
contain **no full-service branch at all** and are in the file purely on
limited-service presence ({zero_fs_txt}). Those rows carry
`no_full_service_office:limited_service_presence_only` in `data_flags` and
necessarily show 0 deposits; read their growth tier as "this is a market the bank
has a foothold in", not "this is a deposit market the bank competes in".

Offices are mapped to markets by county FIPS through the 2023 CBSA crosswalk
(`data/cbsa_crosswalk.csv`), not by the CBSA field on the office record - the
office file carries stale delineations for 5 offices. Three of them (DORA BRANCH,
JASPER BRANCH and MEDICAL CENTER BRANCH, all in Walker County AL, STCNTY 01127)
are coded in the file to CBSA_NO 27530 "Jasper, AL", the micropolitan area that
the 2023 delineations retired when Walker County was added to CBSA 13820
Birmingham, AL; the crosswalk therefore books them to 13820. The other two are
MANNING BRANCH (Clarendon County SC, coded 44940 Sumter, SC -> non-CBSA rural
under the 2023 delineations) and CAMDEN FINANCIAL BANKING CENTER BRANCH (Camden
County GA, coded 41220 St. Marys, GA -> 28680 Kingsland, GA). All five overrides
are recorded in the `flags` column of `data/footprint.csv`.

Total markets: **{len(df)}** ({sizes.get('Major Metro',0)} Major Metro,
{sizes.get('Secondary Metro',0)} Secondary Metro, {sizes.get('Small Metro',0)} Small Metro,
{sizes.get('Micropolitan',0)} Micropolitan, {sizes.get('Rural',0)} Rural county).

## 2. Sources, vintages, pull date

All data pulled 2026-08-16 and cached under `data/` and `raw/`.

| Input | Source | Vintage / series | URL |
|---|---|---|---|
| Population, domestic + net migration, births, deaths | Census Bureau Population Estimates Program | Vintage 2025 (`cbsa-est2025-alldata.csv`, `co-est2025-alldata.csv`); 2020 base and 2025 estimate; migration components cumulated over estimate years 2021-2025 | https://www2.census.gov/programs-surveys/popest/datasets/2020-2025/metro/totals/ and .../counties/totals/ |
| Real GDP by county | BEA Regional, table CAGDP9 (real GDP, chained 2017 dollars) | 2019-2024, released Dec-2025 | https://apps.bea.gov/api/data (LineCode=1, GeoFips=COUNTY) |
| Personal income / population for PCPI | BEA Regional, table CAINC1 (LineCode 1 = personal income, 2 = population) | 2019-2024 | https://apps.bea.gov/api/data |
| Employment | BLS Quarterly Census of Employment and Wages, annual averages, all ownerships, total covered (`own_code=0`, `industry_code=10`) | 2019 and 2024 | https://data.bls.gov/cew/data/api/{{year}}/a/area/{{area}}.csv |
| Deposits | FDIC Summary of Deposits, certs 35583 (Pinnacle Bank) and 873 (Synovus Bank) | June-30 2020 and June-30 2025 | https://banks.data.fdic.gov/api/sod |
| CBSA delineations | OMB / Census delineation file `list1_2023.xlsx` | July 2023 | https://www2.census.gov/programs-surveys/metro-micro/geographies/reference-files/2023/delineation-files/ |

**BEA has no MSA-level GDP or per-capita-income table.** BEA Regional retired the
MAGDP*/MAINC* tables, and `GeoFips=MSA|MIC` is rejected by CAGDP9/CAINC1. MSA and
micropolitan GDP and PCPI were therefore **built by county aggregation** over the
2023 crosswalk: GDP = sum of county real GDP (an additive approximation to
chained dollars); PCPI = sum(county personal income) x 1000 / sum(county BEA
population), i.e. a true population-weighted rate, not a mean of county rates.
BEA merges 53 census counties/independent cities into 24 combination FIPS
(e.g. 51919 Fairfax + Fairfax City + Falls Church); these were resolved by name
(`data/_bea_combo_map.json`) so that **in the CBSA aggregation** no county is
dropped or double-counted. The combo map is applied only on that aggregation
path; the county-level benchmark distribution merges BEA on raw census FIPS and
does drop those counties - see section 5.

## 3. Size tier (Census Vintage-2025 population)

| Tier | Rule |
|---|---|
| Major Metro | Metropolitan Statistical Area, pop2025 >= 1,000,000 |
| Secondary Metro | Metropolitan Statistical Area, 250,000 - 999,999 |
| Small Metro | Metropolitan Statistical Area, pop2025 < 250,000 |
| Micropolitan | Micropolitan Statistical Area (any population) |
| Rural | non-CBSA county |

## 4. Growth indicators

1. **pop_cagr** = (pop2025 / pop2020)^(1/5) - 1. Census PEP Vintage 2025.
2. **netmig_rate** = cumulative net DOMESTIC migration 2021-2025 / pop2020.
   Domestic only, so it measures where Americans are moving between US areas and
   is not contaminated by the post-2021 surge in international arrivals. (At the
   national level domestic migration nets to zero by construction, so the US
   reference row uses NET migration, {pc(natl['netmig_rate'])}, and is shown for
   context only - it is not the centering constant for any z-score.)
3. **gdp_cagr** = (real GDP 2024 / real GDP 2019)^(1/5) - 1, chained 2017 dollars.
4. **emp_cagr** = (QCEW annual-average employment 2024 / 2019)^(1/5) - 1.
5. **pcpi_cagr** = (PCPI 2024 / PCPI 2019)^(1/5) - 1. **Nominal** dollars: BEA
   publishes no county/MSA price deflator. Because every area is scored against
   the same national cross-section over the same window, the common inflation
   component is absorbed by the mean of the distribution and drops out of the
   z-score. Only the *relative* income trajectory survives, which is what the
   composite is meant to capture. The level of pcpi_cagr should not be read as
   real income growth.

US reference CAGRs: population {pc(natl['pop_cagr'])}, real GDP {pc(natl['gdp_cagr'])},
PCPI (nominal) {pc(natl['pcpi_cagr'])}, QCEW employment {pc(natl['emp_cagr'])}.

## 5. Standardization

For each indicator the benchmark distribution is computed across **all 925 US
metropolitan + micropolitan CBSAs**, unweighted (every CBSA counts once, so
Nashville and Aberdeen SD carry equal weight in defining "normal"). The
distribution is winsorized at the 1st and 99th percentiles, and the mean and
population standard deviation of the *winsorized* series are used as the
centering and scaling constants:

    z_i(area) = (x_i(area) - mean_w,i) / std_w,i

The area's own value is not clipped, so a genuine outlier keeps its extreme
z-score; winsorizing only prevents a handful of extremes from setting the scale.

Rural-county markets are scored against the **all-US-county** distribution, built
the same way. A rural county is therefore judged against counties, not against
metros. The county universe has {od['county_rows']:,} rows and the benchmark n
actually used is **{cnty_n['pop_cagr']:,} counties for population,
{cnty_n['netmig_rate']:,} for net migration, {cnty_n['gdp_cagr']:,} for GDP and
{cnty_n['pcpi_cagr']:,} for PCPI** (these are the `n_areas` values on the
`National_Benchmarks` sheet). The {od['county_nan']['gdp_cagr']} counties missing
a GDP/PCPI growth rate are dropped from those two moments, and they are dropped
for identifiable reasons, not silently:

* 51 Virginia counties and independent cities that BEA publishes only inside a
  combination FIPS (51919 etc.). `_bea_combo_map.json` resolves those on the CBSA
  aggregation path but is **not** applied here, so they merge to nothing on raw
  census FIPS. 2 Hawaii counties (Kalawao, Maui) drop for the same reason.
* 9 Connecticut planning regions, whose 2019 GDP is 0 in the BEA extract because
  of the CT county/planning-region break (see Limitations).
* 2 Alaska census areas (Chugach, Copper River) created by the 2019 split of the
  former Valdez-Cordova Census Area, which have no 2019 BEA value.

None of these are footprint markets, and the effect on the moments is
second-order, but the county GDP/PCPI benchmark is a
{cnty_n['gdp_cagr']:,}-county distribution and should be cited as such. Note also
that the row counts of the BEA source files ({od['county_rows']:,}-row county
universe vs. differing BEA file lengths) are *not* the benchmark n; only the
`n_areas` column is.

**emp_cagr limitation.** QCEW was pulled only for footprint areas, so no national
CBSA distribution exists for employment. Its z-score uses the US total-covered
employment CAGR as the center and the winsorized standard deviation of the
{emp_mom[2]} footprint areas as the scale:

    z_emp = (emp_cagr(area) - {natl['emp_cagr']:.6f}) / {emp_mom[1]:.6f}

The footprint is a Southeast-weighted, faster-growing sample than the nation, so
this spread is narrower than a true national CBSA spread would be, which slightly
*inflates* the magnitude of z_emp. emp_cagr carries only a 10% weight partly for
this reason, and rural counties use the same construction.

## 6. Composite and growth tier

    composite_z = 0.225*z_pop + 0.225*z_netmig + 0.225*z_gdp + 0.225*z_pcpi + 0.10*z_emp

The four nationally benchmarked indicators split 90% evenly; employment takes the
remaining 10% because of the weaker benchmark above. When an indicator is missing
for a market, its weight is dropped and the remaining weights are renormalized to
sum to 1; the renormalization is recorded in `data_flags`.

| Growth tier | Rule |
|---|---|
| High | composite_z >= +0.33 |
| Medium | -0.33 < composite_z < +0.33 |
| Low | composite_z <= -0.33 |

+/-0.33 sigma is a deliberately modest cut: on a normal distribution it puts
roughly the top 37% in High and the bottom 37% in Low, so "Medium" means
genuinely middle-of-the-pack rather than merely un-extreme. Result:
{tiers.get('High',0)} High, {tiers.get('Medium',0)} Medium, {tiers.get('Low',0)} Low.

## 7. Merger branches and deposit matching

`branches_acquired_2026` counts offices that entered the footprint through the
January-2026 Synovus acquisition (250 of {od['n_offices']} offices). Those offices are
included in every market's branch and deposit totals: the question this engine
answers is "how attractive is the market the combined bank now sits in", not
"where did the branch come from".

Deposits are matched office-to-office on FDIC UNINUM against the June-2025 and
June-2020 SOD files for both certs. {od['n_offices'] - od['n_no_sod']} of
{od['n_offices']} offices match a 2025 SOD record; {od['n_no_sod']} do not and
carry `no_sod2025_match` in `data/footprint.csv`. The dominant reason is **office
type, not timing**: {od['no_sod_limited']} of the {od['n_no_sod']} are
limited-service offices (loan production, messenger, retail, other) that take no
deposits and never file an SOD record at all. Only {od['no_sod_full']} are
full-service, and those {od['no_sod_full']} are exactly the offices established
after the June-30-2025 survey date - the timing explanation covers
{od['no_sod_new']} of the {od['n_no_sod']}, no more. {od['no_sod_pre2020']} of
the unmatched offices have in fact been open since before 2020. Unmatched offices
contribute 0 to their market's deposit total, so `office_file_deposits_2025_kusd` is
a **lower bound** wherever `pinnacle_limited_service_offices` is non-zero, and is
0 by construction in the markets flagged
`no_full_service_office:limited_service_presence_only`. Separately, 13 June-2025
SOD records for these certs match no office in the August-2026 file (branches
closed or divested between the survey and the file date); their deposits are
excluded from every market total.

**Two deposit columns, two different bases - read the header.** The sheet carries
`office_file_deposits_2025_kusd` (section 1 basis: the June-2025 DEPSUMBR of the
offices present in the **August-2026 office file**, so an office that closed
between the survey and the file date, or that the file cannot match, contributes
nothing) and `proforma_deposits_2025_kusd` (section 8 basis: **every** June-2025
SOD record filed by certs 35583 and 873 in the market's counties). They are not
interchangeable and they do not tie: column totals are
**${_of_tot:,}k vs ${_pf_tot:,}k, a ${_of_gap:,}k gap**, concentrated in
{_of_n} markets. The largest single case is **Bowling Green KY (14540): $0 against
$172,876k** - both Bowling Green offices in the file carry `no_sod2025_match` (one
is an LPO, the other was established 25-Aug-2025, after the survey), while the SOD
itself holds one cert-35583 office with $172,876k in Warren County. The other
divergences are {_of_list}. Every franchise metric in section 8 -
share, BEI, share delta - uses the **pro-forma** column; the office-file column is
retained only because sections 1-7 are built on the office file and it is the
figure that reconciles to `pinnacle_offices`. It was previously named
`pinnacle_deposits_2025_kusd`, which invited exactly the comparison it fails.

`deposit_growth_context` is deliberately a *context string*, not a scored
indicator: it reports the same-branch 2020->2025 deposit CAGR computed over
branches that report **positive deposits at both endpoints**, plus the count of
branches with no 2020 record. Branch-level deposits are booked to the branch of
record and shift with internal reallocations: {od['realloc']} office(s) nationwide
report deposits in June-2020 and a literal $0 in June-2025 (Nashville's main
office, UNINUM 82012, $8,269,186k -> $0, and Albany GA's Five Points branch,
UNINUM 250567, $32,440k -> $0). Those books moved to another office of record; the
deposits were not lost.

**The endpoint rule is now symmetric on both sides of the move, which it was not
before.** Dropping the SOURCE office from the 2020 base while leaving the
RECIPIENT office in the 2025 numerator is not a same-branch comparison - it is a
one-sided one, and it inflated rather than corrected the figure. Nashville printed
a "same-branch CAGR" of **32.4%** because UNINUM 82012's $8.27bn was removed from
the 2020 base while NASHVILLE YARDS (UNINUM 465213, $42,947k -> $13,659,117k), the
office that received the book, stayed in the 2025 numerator. The rule now **pairs**
the two: when an office in the same market shows an absolute increase at least as
large as the whole vanished book, the source's 2020 balance is netted back into
the base, so source and recipient are compared as a single continuing unit and
nothing is discarded. Nashville falls to **8.5%** and Albany GA from 3.8% to
**1.6%**; the pairing is named inline in the context string. Where no office in the
market absorbed a matching amount, the book left the market and the string says so
explicitly instead of implying the problem was handled. This is also why deposit
growth is not reliable enough to enter the composite.

## 8. Limitations

* **Chained-dollar additivity.** Summing county chained-2017-dollar GDP to CBSA
  level is an approximation; BEA warns chained dollars are not strictly additive.
  The bias is second-order over a 5-year window and affects all areas alike.
* **Employment benchmark** - see section 5.
* **Nominal PCPI** - see section 4.
* **Connecticut geography break.** BEA reports old CT counties for 2019-2023 and
  planning regions for 2024, so 7 CT metros have no computable 2019-2024 GDP/PCPI
  growth and are NULL in the national distribution (they are excluded from the
  moments, not zero-filled). No footprint market is affected.
* **Puerto Rico** is absent from the Census and BEA extracts, so the national
  universe is the 50 states + DC. No footprint market is affected.
* **CBSA delineation drift.** The office file's own CBSA fields are pre-2023 in
  places (see section 1): the three Walker County AL offices are still coded to
  27530 "Jasper, AL", the micropolitan area the 2023 delineations dissolved into
  Birmingham. The 2023 vintage also retired the "Birmingham-Hoover" title, so the
  market is now CBSA 13820 "Birmingham, AL" with a Vintage-2025 population of
  1,197,766, i.e. above the 1,000,000 Major Metro threshold. Anyone expecting a
  "Secondary Metro" label for Birmingham is working from a pre-2023 delineation
  or a divisional population.
* **Single-vintage endpoints.** All CAGRs use two endpoints; a COVID-distorted
  2020 or 2019 base is not smoothed. 2019->2024 was chosen for the economic
  series precisely to bracket rather than start inside the pandemic.
* **QCEW micro coverage.** Micropolitan CBSAs resolve at QCEW aggregation level
  80, not 40. Kingsland GA (28680) had no 2019 micro-area file; it is a
  single-county area (Camden County GA) whose 2024 micro total equals the county
  total exactly, so 2019 was backfilled from county 13039.
"""
    if realloc is not None and "franchise_stage" in df.columns:
        txt += franchise_methodology(df, realloc)
    with open(os.path.join(OUT, "methodology.md"), "w", encoding="utf-8") as fh:
        fh.write(txt)


# =========================================================================== #
# =========================================================================== #
# ==                                                                       == #
# ==  MODULE 10 -- FRANCHISE STAGE DIMENSION                               == #
# ==  (appended 2026-08-17; does NOT touch the size/growth scoring above)  == #
# ==                                                                       == #
# ==  Question answered: "where is newco Pinnacle on its journey in this   == #
# ==  market?"  Every metric below is PRO-FORMA NEWCO: FDIC certs 35583    == #
# ==  (legacy Pinnacle Bank) and 873 (Synovus Bank) are summed in EVERY    == #
# ==  year, because the January-2026 merger means the combined company's   == #
# ==  history in a market is the EARLIER of the two predecessors'.         == #
# ==                                                                       == #
# =========================================================================== #
# =========================================================================== #

# --- franchise-stage configuration (every threshold is named here) ---------- #
SOD_CURRENT_YEAR = 2025          # latest June-30 SOD vintage available
SOD_BASE_YEAR = 2020             # 5-year comparison base
ENTRENCHED_DEPOSIT_SHARE = 0.08  # 8% deposit share -> 'Entrenched Leader'
NEW_ENTRANT_ENTRY_YEAR = 2021    # entry_year >= this -> 'New Entrant'
NEW_ENTRANT_TENURE = 8           # years
NEW_ENTRANT_MAX_BRANCHES = 2
NEW_ENTRANT_MAX_SHARE = 0.02     # 2% deposit share
PULLBACK_NET_BRANCHES = -2       # net_branch_change_5yr <= this
PULLBACK_BRANCH_PCT = -0.25      # branch_change_pct <= this (-25%)
EXPANDING_NET_BRANCHES = 2       # net_branch_change_5yr >= this
EXPANDING_SHARE_BPS = 50.0       # share_delta_5yr_bps >= this
EXPANDING_YOUNG_TENURE = 12      # years
PROFORMA_CERTS = (35583, 873)
# A market whose pro-forma deposits fall by more than this share of the 2020
# book is checked for a main-office reallocation wart (DEPSUMBR moved to another
# office of record).  NOTE (2026-08-17 fix): this test is applied REGARDLESS of
# the branch trajectory.  The previous version exempted markets with
# net_branch_change_5yr <= -2, which switched the check off for the two largest
# deposit collapses in the book (Columbus GA-AL and Miami) -- a bank that closes
# branches AND loses a third of its book is the LEAST plausible case to exempt
# from a booking-artifact check, not the most.
REALLOC_DROP_FLAG = -0.20
# Office-level book relocation detector (section 8.6).  An office whose SOD
# DEPSUMBR moves by at least this many $k between the two endpoint surveys is a
# candidate book relocation rather than organic growth/attrition.
RELOC_OFFICE_MIN_KUSD = 1_000_000     # $1.0bn
# ... or, in a smaller market, an office whose absolute move is at least this
# multiple of the market's own 2020 pro-forma book (subject to the floor below,
# so a tiny market cannot generate a "relocation" out of ordinary growth).
RELOC_OFFICE_MIN_FRAC = 0.25
RELOC_OFFICE_FLOOR_KUSD = 25_000      # $25m: below this, nothing is material
# Size alone is not evidence of a booking move -- a big branch growing fast is
# still a branch. The office must ALSO look anomalous in its own terms:
RELOC_SOURCE_RETAINED = 0.20          # source kept <=20% of its own 2020 book
RELOC_RECIPIENT_MULT = 5.0            # recipient ended at >=5x its own 2020 book
# (an office absent in 2020 counts as a recipient only on the absolute $1bn test)


def _county_market_maps():
    """
    Two county -> market-key maps.

    strict : the 151 counties in which newco actually operates an office today.
             This is EXACTLY the county set covered by data/sod_market_county.csv
             (the all-institution denominator), so every share computed off it is
             numerator/denominator consistent -- same counties on both sides.
    wide   : strict, plus any other county that belongs to one of the 81 footprint
             CBSAs via the 2023 crosswalk. Used ONLY for entry_year, so that a
             1990s/2000s branch in a CBSA county the bank has since exited still
             counts as "we were in this market then".
    """
    fp = pd.read_csv(p("footprint.csv"), dtype={"county_fips": str, "cbsa_code": str})
    fp["county_fips"] = fp["county_fips"].str.zfill(5)
    fp["key"] = np.where(fp["cbsa_code"].eq("RURAL"), fp["county_fips"], fp["cbsa_code"])
    strict = dict(zip(fp["county_fips"], fp["key"]))

    xw = pd.read_csv(p("cbsa_crosswalk.csv"), dtype={"county_fips": str, "cbsa_code": str})
    xw["county_fips"] = xw["county_fips"].str.zfill(5)
    cbsa_keys = set(strict.values())
    wide = dict(strict)
    for cnty, cbsa in zip(xw["county_fips"], xw["cbsa_code"]):
        if cnty not in wide and cbsa in cbsa_keys:
            wide[cnty] = cbsa
    return strict, wide


def deposit_taking_counties() -> set:
    """
    The subset of the 151 footprint counties in which newco can actually HOLD a
    deposit share: at least one FULL-SERVICE office today, or pro-forma SOD
    deposits in either endpoint survey.

    Why this exists (2026-08-17 fix). `strict` is built from every row of
    footprint.csv, including counties where the only presence is a
    loan-production / limited-service office. An LPO cannot file a Summary of
    Deposits, so those counties contribute a structurally ZERO numerator while
    still contributing their full all-institution denominator -- the share is
    mechanically understated. Nine counties are LPO-only; five of them ARE their
    whole market (those markets are 'Limited-Service Foothold' and get a null
    share), but four sit inside full-service markets and silently diluted them:

        51680 Lynchburg city VA -> 31340 Lynchburg VA        (54.5% of denominator)
        51067 Franklin VA       -> 40220 Roanoke VA          (15.2%)
        28033 DeSoto MS         -> 32820 Memphis TN-MS-AR    (12.2%)
        47167 Tipton TN         -> 32820 Memphis TN-MS-AR    ( 2.7%)
        45091 York SC           -> 16740 Charlotte NC-SC     ( 1.0%)

    Confirmed against sod_history.csv: pro-forma deposits in all nine counties
    are $0 in both 2020 and 2025, so removing them from the denominator removes
    nothing from the numerator and the share stays numerator/denominator
    consistent -- now over the counties where a share is definable.
    """
    br = load_branches()
    fs = set(br.loc[br["full_service"], "county_fips"])
    hist = load_sod_history()
    dep = hist[(hist["year"].isin([SOD_BASE_YEAR, SOD_CURRENT_YEAR]))
               & (hist["deposits_kusd"] > 0)]
    return fs | set(dep["county_fips"])


def load_sod_history():
    """
    data/sod_history.csv -- one row per cert-year-county, 1994-2025, certs 35583
    and 873. Columns: cert, year, county_fips, branches, deposits_kusd.

    Caveat carried from the fetch (documented in methodology section 8): cert 873
    before 2010 is only the legacy Columbus Bank & Trust charter, not the whole
    Synovus franchise -- the other Synovus banks sat under sibling certs. Pre-2010
    branch counts are therefore a FLOOR, and an entry_year derived from them is a
    LATEST-POSSIBLE entry date, never too early.
    """
    h = pd.read_csv(p("sod_history.csv"), dtype={"county_fips": str})
    h["county_fips"] = h["county_fips"].str.zfill(5)
    h = h[h["cert"].isin(PROFORMA_CERTS)]
    for c in ("year", "branches", "deposits_kusd"):
        h[c] = pd.to_numeric(h[c], errors="coerce")
    return h.dropna(subset=["year"])


def load_sod_history_office():
    """
    data/sod_history_office.csv -- one row per cert-year-OFFICE (UNINUM), 1994-2025,
    certs 35583 and 873, extracted from the same raw/sod_hist/*.json payloads that
    produced sod_history.csv. Columns: cert, year, uninum, county_fips, deposits_kusd.

    Deposit totals reconcile to sod_history.csv exactly for every cert-year-county
    in 2020 and 2025. Needed because a book relocation is an OFFICE-level event:
    it is invisible in a county aggregate when both offices sit in the same county
    (Nashville) and looks like organic decline when they do not (Columbus GA ->
    Atlanta).
    """
    o = pd.read_csv(p("sod_history_office.csv"),
                    dtype={"county_fips": str, "uninum": str})
    o["county_fips"] = o["county_fips"].str.zfill(5)
    o = o[o["cert"].isin(PROFORMA_CERTS)]
    for c in ("year", "deposits_kusd"):
        o[c] = pd.to_numeric(o[c], errors="coerce")
    return o.dropna(subset=["year"])


def book_relocation_moves():
    """
    Office-level book-relocation detector (section 8.6).

    For every pro-forma SOD office, compare DEPSUMBR at the two endpoint surveys
    ({SOD_BASE_YEAR} vs {SOD_CURRENT_YEAR}); an office absent from a vintage
    contributes 0 for that vintage. An office whose absolute move clears
    RELOC_OFFICE_MIN_KUSD ($1.0bn) OR RELOC_OFFICE_MIN_FRAC (25%) of its market's
    own 2020 pro-forma book is reported as a MATERIAL BOOK MOVEMENT -- a candidate
    relocation of a book of record rather than organic gathering or attrition.

    This is a diagnostic, not a stage input: it changes no label. Its job is to
    make the movement visible on the row and in the methodology, which is exactly
    what the previous version failed to do for the Columbus GA -> Atlanta case.

    Returns {market_key: {out_kusd, in_kusd, offices: [...]}} plus a "_pairs" entry
    listing the largest source/recipient candidates across markets.
    """
    o = load_sod_history_office()
    strict, wide = _county_market_maps()
    o = o.assign(key=o["county_fips"].map(wide)).dropna(subset=["key"])
    ends = o[o["year"].isin([SOD_BASE_YEAR, SOD_CURRENT_YEAR])]
    piv = (ends.pivot_table(index=["uninum", "cert", "key"], columns="year",
                            values="deposits_kusd", aggfunc="sum")
           .reindex(columns=[SOD_BASE_YEAR, SOD_CURRENT_YEAR]).fillna(0.0))
    piv["delta"] = piv[SOD_CURRENT_YEAR] - piv[SOD_BASE_YEAR]
    piv = piv.reset_index()

    base = (o[o["year"] == SOD_BASE_YEAR].groupby("key")["deposits_kusd"].sum()
            .to_dict())
    mkt_base = piv["key"].map(lambda k: float(base.get(k, 0.0)))
    d20, d25, ad = piv[SOD_BASE_YEAR], piv[SOD_CURRENT_YEAR], piv["delta"].abs()

    # (a) material in dollar terms
    material = (ad >= RELOC_OFFICE_MIN_KUSD) | (
        (ad >= RELOC_OFFICE_MIN_FRAC * mkt_base) & (ad >= RELOC_OFFICE_FLOOR_KUSD))
    # (b) anomalous in the office's own terms -- this is what separates a book
    #     relocation from a large branch simply growing or shrinking.
    is_source = (d20 > 0) & (d25 <= RELOC_SOURCE_RETAINED * d20)
    is_recipient = (((d20 > 0) & (d25 >= RELOC_RECIPIENT_MULT * d20))
                    | ((d20 <= 0) & (piv["delta"] >= RELOC_OFFICE_MIN_KUSD)))
    mat = piv[material & (is_source | is_recipient)].copy()

    names = {}
    try:
        br = load_branches()
        names = dict(zip(br["uninum"].astype(str), br["offname"].astype(str)))
    except Exception:
        pass

    out = {}
    for key, g in mat.groupby("key"):
        recs = []
        for _, r in g.sort_values("delta").iterrows():
            recs.append({
                "uninum": str(r["uninum"]), "cert": int(r["cert"]),
                "offname": names.get(str(r["uninum"]),
                                     "(office closed / not in current office file)"),
                f"dep_{SOD_BASE_YEAR}_kusd": int(r[SOD_BASE_YEAR]),
                f"dep_{SOD_CURRENT_YEAR}_kusd": int(r[SOD_CURRENT_YEAR]),
                "delta_kusd": int(r["delta"]),
            })
        out[key] = {
            "out_kusd": int(-g.loc[g["delta"] < 0, "delta"].sum()),
            "in_kusd": int(g.loc[g["delta"] > 0, "delta"].sum()),
            "offices": recs,
        }
    return out


def load_market_totals():
    """
    data/sod_market_county.csv -- ALL-INSTITUTION June-30 SOD county totals for
    2020 and 2025 across the 151 footprint counties.
    Columns: county_fips, year, market_deposits_kusd, market_branches, n_institutions.
    """
    m = pd.read_csv(p("sod_market_county.csv"), dtype={"county_fips": str})
    m["county_fips"] = m["county_fips"].str.zfill(5)
    return m


def build_franchise(br: pd.DataFrame) -> pd.DataFrame:
    """
    One row per market key with the pro-forma franchise metrics.

    Numerator (newco) and denominator (all institutions) are aggregated over the
    SAME county set in the same year, so a share is a true within-market share.
    """
    strict, wide = _county_market_maps()
    hist = load_sod_history()
    mkt = load_market_totals()

    hist["key_strict"] = hist["county_fips"].map(strict)
    hist["key_wide"] = hist["county_fips"].map(wide)

    # ---- entry_year: earliest SOD year with >=1 pro-forma branch (wide map) ---
    ent = hist[(hist["branches"] >= 1) & hist["key_wide"].notna()]
    entry = ent.groupby("key_wide")["year"].min().astype(int).to_dict()

    # ---- pro-forma deposits, strict county set (share numerator) -------------
    hs = hist[hist["key_strict"].notna()]
    pf = (hs[hs["year"].isin([SOD_BASE_YEAR, SOD_CURRENT_YEAR])]
          .groupby(["key_strict", "year"])[["branches", "deposits_kusd"]].sum())

    # ---- pro-forma BRANCH COUNTS, wide county set ----------------------------
    # 2026-08-17 fix. Branch counts are a bank-side census of "how many offices do
    # we run in this market", not a share, so they must cover the whole CBSA, not
    # only the counties the bank still occupies today. Computing branches_2020 off
    # the strict map silently understated it by one branch in three markets --
    # 36740 Orlando (Lake Cty FL 12069, $16,849k), 45300 Tampa (Hernando 12053,
    # $81,880k) and 24860 Greenville SC (Anderson 45007, $27,251k) -- because the
    # bank held an office there in 2020 in a CBSA county it has since exited.
    # That inflated net_branch_change_5yr by +1 in each and flipped Orlando to
    # 'Expanding' on the young-market clause alone. No 2025 SOD branch sits in a
    # wide-only county, so branches_2025_sod is identical on both maps and
    # branch_share_2025 remains numerator/denominator consistent.
    hw = hist[hist["key_wide"].notna()]
    pfb = (hw[hw["year"].isin([SOD_BASE_YEAR, SOD_CURRENT_YEAR])]
           .groupby(["key_wide", "year"])[["branches"]].sum())
    wide_only = hw[hw["key_strict"].isna() & (hw["branches"] >= 1)
                   & hw["year"].isin([SOD_BASE_YEAR, SOD_CURRENT_YEAR])]
    wide_only_keys = {}
    for _, r in wide_only.iterrows():
        wide_only_keys.setdefault(r["key_wide"], []).append(
            f"{r['county_fips']}@{int(r['year'])}:{int(r['branches'])}br")

    # ---- all-institution market totals, deposit-taking counties only ---------
    dep_counties = deposit_taking_counties()
    mkt = mkt[mkt["county_fips"].isin(dep_counties)]
    mkt = mkt.assign(key=mkt["county_fips"].map(strict)).dropna(subset=["key"])
    mt = mkt.groupby(["key", "year"])[
        ["market_deposits_kusd", "market_branches", "n_institutions"]].sum()
    # n_institutions is a per-county distinct count; summing it over a multi-county
    # market double-counts banks present in two counties. It is reported as
    # "institution-county presences" for multi-county markets -- see methodology.
    n_counties = mkt.groupby("key")["county_fips"].nunique().to_dict()

    # ---- current office mix / merger provenance from the office file ---------
    br = br.copy()
    fs = br[br["full_service"]]
    prov = {}
    for key, grp in br.groupby("key"):
        g = fs[fs["key"] == key]
        basis = g if len(g) else grp          # LPO-only markets fall back to all offices
        prov[key] = bool(basis["acquired_2026"].all())

    def g(idx, col, default=0.0):
        try:
            v = pf.loc[idx, col]
        except KeyError:
            return default
        return float(v) if pd.notna(v) else default

    def gm(idx, col, default=float("nan")):
        try:
            v = mt.loc[idx, col]
        except KeyError:
            return default
        return float(v) if pd.notna(v) else default

    def gb(idx, default=0.0):
        try:
            v = pfb.loc[idx, "branches"]
        except KeyError:
            return default
        return float(v) if pd.notna(v) else default

    reloc = book_relocation_moves()

    rows = []
    for key in sorted(set(br["key"])):
        b25 = int(gb((key, SOD_CURRENT_YEAR)))
        b20 = int(gb((key, SOD_BASE_YEAR)))
        d25 = g((key, SOD_CURRENT_YEAR), "deposits_kusd")
        d20 = g((key, SOD_BASE_YEAR), "deposits_kusd")
        md25 = gm((key, SOD_CURRENT_YEAR), "market_deposits_kusd")
        md20 = gm((key, SOD_BASE_YEAR), "market_deposits_kusd")
        mb25 = gm((key, SOD_CURRENT_YEAR), "market_branches")
        ni25 = gm((key, SOD_CURRENT_YEAR), "n_institutions")

        s25 = d25 / md25 if md25 and np.isfinite(md25) and md25 > 0 else None
        s20 = d20 / md20 if md20 and np.isfinite(md20) and md20 > 0 else None
        bps = round((s25 - s20) * 10000, 1) if (s25 is not None and s20 is not None) else None
        bshare = b25 / mb25 if mb25 and np.isfinite(mb25) and mb25 > 0 else None
        bei = (s25 / bshare) if (bshare not in (None, 0) and s25 is not None) else None

        ey = entry.get(key)
        rows.append({
            "key": key,
            "entry_year": ey,
            "tenure_years": (SOD_CURRENT_YEAR - ey) if ey else None,
            "branches_2020": b20,
            "branches_2025_sod": b25,
            "net_branch_change_5yr": b25 - b20,
            "branch_change_pct": ((b25 / b20) - 1) if b20 else None,
            "proforma_deposits_2020_kusd": int(d20),
            "proforma_deposits_2025_kusd": int(d25),
            "market_deposits_2020_kusd": int(md20) if np.isfinite(md20) else None,
            "market_deposits_2025_kusd": int(md25) if np.isfinite(md25) else None,
            "market_branches_2025": int(mb25) if np.isfinite(mb25) else None,
            "n_institutions": int(ni25) if np.isfinite(ni25) else None,
            "n_market_counties": n_counties.get(key, 0),
            "deposit_share_2020": s20,
            "deposit_share_2025": s25,
            "share_delta_5yr_bps": bps,
            "branch_share_2025": bshare,
            "branch_effectiveness_index": bei,
            "entered_via_2026_merger": prov.get(key, False),
            "book_moved_out_kusd": int(reloc.get(key, {}).get("out_kusd", 0)),
            "book_moved_in_kusd": int(reloc.get(key, {}).get("in_kusd", 0)),
            "branches_2020_wide_only": "|".join(wide_only_keys.get(key, [])),
        })
    return pd.DataFrame(rows).set_index("key")


# --------------------------------------------------------------------------- #
# 10b. The stage rules -- deterministic, ordered, first match wins
# --------------------------------------------------------------------------- #
STAGE_DEFS = [
    ("Limited-Service Foothold",
     "Zero FULL-SERVICE branches today: presence is a loan-production / limited-service "
     "office only, so the bank files no Summary of Deposits record and holds no measurable "
     "deposit share. Checked first because every deposit-based rule below is undefined here."),
    ("New Entrant",
     f"entry_year >= {NEW_ENTRANT_ENTRY_YEAR} (first pro-forma SOD branch in the last 5 "
     f"surveys) OR (tenure < {NEW_ENTRANT_TENURE} yrs AND branches_2025 <= "
     f"{NEW_ENTRANT_MAX_BRANCHES} AND deposit_share_2025 < "
     f"{NEW_ENTRANT_MAX_SHARE*100:.0f}%). Young and sub-scale. A market with at least one "
     "full-service branch today but NO SOD history at all is also a New Entrant: the office "
     "was established after the June-2025 survey date and has not yet filed a Summary of "
     "Deposits (e.g. Decatur AL, opened 27-Jan-2026)."),
    ("Entrenched Leader (consolidating)",
     f"BOTH tests fire: deposit_share_2025 >= {ENTRENCHED_DEPOSIT_SHARE*100:.0f}% AND the "
     "Pulling-Back branch/share test (rule 4). Evaluated BEFORE plain Pulling Back and "
     "plain Entrenched Leader, so a dominant franchise thinning overlapping branches is "
     "never mislabelled as a retreat."),
    ("Pulling Back",
     f"net_branch_change_5yr <= {PULLBACK_NET_BRANCHES} AND share_delta_5yr_bps < 0; "
     f"OR branch_change_pct <= {PULLBACK_BRANCH_PCT*100:.0f}% with a measured non-positive "
     f"share delta (share_delta_5yr_bps <= 0). Physical and share retreat together."),
    ("Entrenched Leader",
     f"deposit_share_2025 >= {ENTRENCHED_DEPOSIT_SHARE*100:.0f}%. A market where the bank "
     "is a structural incumbent."),
    ("Expanding",
     f"net_branch_change_5yr >= +{EXPANDING_NET_BRANCHES} OR share_delta_5yr_bps >= "
     f"+{EXPANDING_SHARE_BPS:.0f} OR (tenure < {EXPANDING_YOUNG_TENURE} yrs AND "
     "share_delta_5yr_bps > 0 AND branches grew)."),
    ("Established - Mature",
     "Everything else: long presence, share and branch count broadly stable, below the "
     "entrenchment cutoff."),
]
# STAGE_DEFS is printed in TRUE EVALUATION ORDER (2026-08-17 fix). The previous
# version listed 'Entrenched Leader (consolidating)' last, below the
# 'Everything else' catch-all, in a table whose entire premise is order.


def classify_stage(r) -> str:
    """Ordered rules; the FIRST match wins. `r` is a franchise-metrics mapping."""
    n_full = r["pinnacle_full_service_branches"]
    ey, ten = r["entry_year"], r["tenure_years"]
    b25, b20 = r["branches_2025_sod"], r["branches_2020"]
    net, pct = r["net_branch_change_5yr"], r["branch_change_pct"]
    s25, bps = r["deposit_share_2025"], r["share_delta_5yr_bps"]

    # 1 ------------------------------------------------------------------ #
    if not n_full:
        return "Limited-Service Foothold"

    # 2 ------------------------------------------------------------------ #
    if ey is None:
        # Full-service branch today, zero SOD history: opened after the June-2025
        # survey, so it cannot yet have an entry_year or a share.
        return "New Entrant"
    if ey >= NEW_ENTRANT_ENTRY_YEAR:
        return "New Entrant"
    if (ten is not None and ten < NEW_ENTRANT_TENURE
            and b25 <= NEW_ENTRANT_MAX_BRANCHES
            and s25 is not None and s25 < NEW_ENTRANT_MAX_SHARE):
        return "New Entrant"

    # 3 / 4 --------------------------------------------------------------- #
    # 2026-08-17 fix: the second clause previously read `(bps is None or bps <= 0)`,
    # i.e. an UNDEFINED share delta was treated as satisfying "no share gain" --
    # which is not what the printed rule says. The executed rule is now the printed
    # rule: a share gain must be measured and non-positive. (No market took the
    # bps-is-None path in this run, so no label moves.)
    pulling = ((net <= PULLBACK_NET_BRANCHES and bps is not None and bps < 0)
               or (pct is not None and pct <= PULLBACK_BRANCH_PCT
                   and bps is not None and bps <= 0))
    entrenched = s25 is not None and s25 >= ENTRENCHED_DEPOSIT_SHARE
    if pulling and entrenched:
        return "Entrenched Leader (consolidating)"   # documented borderline rule
    if pulling:
        return "Pulling Back"
    if entrenched:
        return "Entrenched Leader"

    # 5 ------------------------------------------------------------------- #
    if net >= EXPANDING_NET_BRANCHES:
        return "Expanding"
    if bps is not None and bps >= EXPANDING_SHARE_BPS:
        return "Expanding"
    if (ten is not None and ten < EXPANDING_YOUNG_TENURE
            and bps is not None and bps > 0 and net > 0):
        return "Expanding"

    # 6 ------------------------------------------------------------------- #
    return "Established - Mature"


def stage_detail(stage: str, r) -> str:
    """'<stage>; share x.x% (+/-Nbps 5yr); n branches (+/-d 5yr); entered YYYY[; via 2026 merger]'"""
    s25 = r["deposit_share_2025"]
    sh = "n/a" if s25 is None else f"{s25*100:.1f}%"
    bps = r["share_delta_5yr_bps"]
    bt = "n/a bps" if bps is None else f"{bps:+.0f}bps"
    net = r["net_branch_change_5yr"]
    ey = r["entry_year"]
    txt = (f"{stage}; share {sh} ({bt} 5yr); {r['branches_2025_sod']} branches "
           f"({net:+d} 5yr); entered {ey if ey else 'n/a'}")
    if r["entered_via_2026_merger"]:
        txt += "; via 2026 merger"
    # 2026-08-17 fix. A share move driven by a book of record changing office is
    # not an operating result. Say so on the row, next to the label it drove.
    mo = int(r.get("book_moved_out_kusd", 0) or 0)
    mi = int(r.get("book_moved_in_kusd", 0) or 0)
    if mo or mi:
        parts = []
        if mo:
            parts.append(f"${mo/1e6:,.1f}bn booked OUT of offices here")
        if mi:
            parts.append(f"${mi/1e6:,.1f}bn booked IN to offices here")
        txt += (f"; CAUTION: {' and '.join(parts)} between the 2020 and 2025 surveys "
                f"(office-level book relocation, see methodology 8.6) -- check whether "
                f"the move was intra-market (share unaffected) or cross-market (share "
                f"delta is partly a booking change, not an operating result) before "
                f"reading this label")
    return txt


def reallocation_check(hist=None) -> dict:
    """
    KNOWN DATA WART. Cert 35583's Nashville main office (UNINUM 82012) reported
    DEPSUMBR = 0 in the June-2025 SOD after an internal book reallocation
    ($8.27bn in June-2020). Share LEVELS keep SOD as-reported -- the money did not
    leave the bank -- but we must establish WHERE the 2025 book sits before
    trusting the Nashville share. The test: compare cert-35583 deposits in
    Davidson County TN (47037) 2020 vs 2025. If the county total GREW, the book
    was reallocated to another Davidson office and the MSA-level share is
    unaffected. Same check for the Five Points branch (UNINUM 250567, cert 873).
    """
    hist = load_sod_history() if hist is None else hist
    out = {}
    dav = hist[(hist["cert"] == 35583) & (hist["county_fips"] == "47037")]
    d20 = float(dav[dav["year"] == 2020]["deposits_kusd"].sum())
    d25 = float(dav[dav["year"] == 2025]["deposits_kusd"].sum())
    b20 = int(dav[dav["year"] == 2020]["branches"].sum())
    b25 = int(dav[dav["year"] == 2025]["branches"].sum())
    out["davidson"] = {
        "cert": 35583, "county": "47037 Davidson TN",
        "dep_2020_kusd": int(d20), "dep_2025_kusd": int(d25),
        "branches_2020": b20, "branches_2025": b25,
        "pct_change": (d25 / d20 - 1) if d20 else None,
        "book_retained_in_county": bool(d20 and d25 >= d20),
    }
    br = load_branches()
    fv = br[br["uninum"] == "250567"]
    if len(fv):
        row = fv.iloc[0]
        cty = row["county_fips"]
        syn = hist[(hist["cert"] == 873) & (hist["county_fips"] == cty)]
        s20 = float(syn[syn["year"] == 2020]["deposits_kusd"].sum())
        s25 = float(syn[syn["year"] == 2025]["deposits_kusd"].sum())
        out["five_points"] = {
            "uninum": "250567", "offname": str(row["offname"]), "cert": 873,
            "county": cty,
            "branch_dep_2020_kusd": None if pd.isna(row["deposits_2020_kusd"])
                                    else int(row["deposits_2020_kusd"]),
            "branch_dep_2025_kusd": None if pd.isna(row["deposits_2025_kusd"])
                                    else int(row["deposits_2025_kusd"]),
            "county_dep_2020_kusd": int(s20), "county_dep_2025_kusd": int(s25),
            "book_retained_in_county": bool(s20 and s25 >= s20),
            "zero_reported_2025": bool(pd.notna(row["deposits_2025_kusd"])
                                       and row["deposits_2025_kusd"] <= 0),
        }
    else:
        out["five_points"] = {"uninum": "250567", "found": False}
    return out


def attach_franchise(df: pd.DataFrame, br: pd.DataFrame):
    """
    Merge the franchise metrics onto the scored market frame, classify the stage,
    and flag any market whose pro-forma deposits fell implausibly against a stable
    branch count (the main-office reallocation wart).
    """
    fr = build_franchise(br)
    df = df.merge(fr, left_on="cbsa_code_or_fips", right_index=True, how="left")

    def _none(v):
        return None if (v is None or (isinstance(v, float) and not np.isfinite(v))) else v

    stages, details, flags = [], [], []
    for _, row in df.iterrows():
        r = {k: _none(row[k]) for k in
             ["entry_year", "tenure_years", "branches_2020", "branches_2025_sod",
              "net_branch_change_5yr", "branch_change_pct", "deposit_share_2025",
              "deposit_share_2020", "share_delta_5yr_bps"]}
        r["entry_year"] = int(r["entry_year"]) if r["entry_year"] is not None else None
        r["tenure_years"] = int(r["tenure_years"]) if r["tenure_years"] is not None else None
        for k in ("branches_2020", "branches_2025_sod", "net_branch_change_5yr"):
            r[k] = int(r[k] or 0)
        r["entered_via_2026_merger"] = bool(row["entered_via_2026_merger"])
        r["pinnacle_full_service_branches"] = int(row["pinnacle_full_service_branches"])
        r["book_moved_out_kusd"] = int(row.get("book_moved_out_kusd", 0) or 0)
        r["book_moved_in_kusd"] = int(row.get("book_moved_in_kusd", 0) or 0)
        st = classify_stage(r)
        stages.append(st)
        details.append(stage_detail(st, r))

        f = [x for x in str(row["data_flags"]).split("|") if x]
        d20, d25 = row["proforma_deposits_2020_kusd"], row["proforma_deposits_2025_kusd"]
        if d20 and np.isfinite(d20) and d20 > 0:
            drop = d25 / d20 - 1
            # 2026-08-17 fix: no branch-count exemption. See REALLOC_DROP_FLAG.
            if drop <= REALLOC_DROP_FLAG:
                f.append(f"deposit_book_drop_{drop*100:.0f}pct:"
                         "check_main_office_reallocation")
        # Office-level book relocation, source and recipient side alike.
        mo, mi = int(row.get("book_moved_out_kusd", 0) or 0), \
                 int(row.get("book_moved_in_kusd", 0) or 0)
        if mo:
            f.append(f"book_relocation_out:${mo/1e6:,.1f}bn_left_offices_in_this_market"
                     ":stage_label_reflects_a_booking_move_not_only_operations")
        if mi:
            f.append(f"book_relocation_in:${mi/1e6:,.1f}bn_arrived_at_offices_in_this_"
                     "market:stage_label_reflects_a_booking_move_not_only_operations")
        wo = str(row.get("branches_2020_wide_only", "") or "")
        if wo:
            f.append(f"branches_2020_includes_cbsa_county_since_exited:{wo}")
        if r["entry_year"] is None:
            f.append("no_sod_history:entry_year_unavailable")
        flags.append("|".join(f))

    df["franchise_stage"] = stages
    df["stage_detail"] = details
    df["data_flags"] = flags

    # Nashville main-office wart: state it on the row itself.
    rc = reallocation_check()
    dav = rc["davidson"]
    nash = df["cbsa_code_or_fips"] == "34980"
    if nash.any():
        note = (f"nashville_main_office_uninum82012_reported_$0_in_june2025_sod;"
                f"cert35583_davidson_county_deposits_2020={dav['dep_2020_kusd']:,}k"
                f"->2025={dav['dep_2025_kusd']:,}k"
                f"({'book_retained_in_county_share_unaffected' if dav['book_retained_in_county'] else 'BOOK_LEFT_COUNTY_SHARE_UNDERSTATED'})")
        df.loc[nash, "data_flags"] = (df.loc[nash, "data_flags"].fillna("") + "|" + note).str.strip("|")
    return df, rc


def franchise_pivot(df: pd.DataFrame):
    """stage x size_tier and stage x growth_tier: market counts + pro-forma deposits."""
    recs = []
    for stage, g in df.groupby("franchise_stage"):
        for dim, col in (("size_tier", "size_tier"), ("growth_tier", "growth_tier")):
            for tier, gg in g.groupby(col):
                recs.append({
                    "franchise_stage": stage, "dimension": dim, "tier": tier,
                    "markets": len(gg),
                    "proforma_deposits_2025_kusd": int(gg["proforma_deposits_2025_kusd"].sum()),
                    "market_deposits_2025_kusd": int(gg["market_deposits_2025_kusd"].fillna(0).sum()),
                    "branches_2025_sod": int(gg["branches_2025_sod"].fillna(0).sum()),
                })
    piv = pd.DataFrame(recs).sort_values(
        ["dimension", "franchise_stage", "tier"]).reset_index(drop=True)
    tot = df.groupby("franchise_stage").agg(
        markets=("market_name", "size"),
        proforma_deposits_2025_kusd=("proforma_deposits_2025_kusd", "sum"),
        market_deposits_2025_kusd=("market_deposits_2025_kusd", lambda s: s.fillna(0).sum()),
        branches_2025_sod=("branches_2025_sod", lambda s: s.fillna(0).sum()),
    ).reset_index()
    tot.insert(1, "dimension", "TOTAL")
    tot.insert(2, "tier", "(all markets)")
    return pd.concat([tot, piv], ignore_index=True)


def write_franchise_sheet(wb, df):
    """Franchise_Stage sheet: the pivot, then the stage definitions."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)
    piv = franchise_pivot(df)

    ws = wb.create_sheet("Franchise_Stage")
    ws.append(list(piv.columns))
    for c in range(1, len(piv.columns) + 1):
        ws.cell(row=1, column=c).fill = hdr_fill
        ws.cell(row=1, column=c).font = hdr_font
    for rec in piv.itertuples(index=False):
        ws.append([None if (isinstance(v, float) and math.isnan(v)) else v for v in rec])
    for i, col in enumerate(piv.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            34 if col == "franchise_stage" else 26 if col.endswith("kusd") else 18
        if col.endswith("kusd") or col in ("markets", "branches_2025_sod"):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = "#,##0"
    ws.freeze_panes = "A2"

    r0 = ws.max_row + 3
    ws.cell(row=r0, column=1, value="FRANCHISE STAGE DEFINITIONS "
            "(ordered rules, first match wins; see methodology.md section 8)").font = Font(bold=True)
    ws.cell(row=r0 + 1, column=1, value="order").font = hdr_font
    ws.cell(row=r0 + 1, column=1).fill = hdr_fill
    ws.cell(row=r0 + 1, column=2, value="franchise_stage").font = hdr_font
    ws.cell(row=r0 + 1, column=2).fill = hdr_fill
    ws.cell(row=r0 + 1, column=3, value="rule").font = hdr_font
    ws.cell(row=r0 + 1, column=3).fill = hdr_fill
    for j, (name, rule) in enumerate(STAGE_DEFS):
        rr = r0 + 2 + j
        ws.cell(row=rr, column=1, value=f"{j+1}")
        ws.cell(row=rr, column=2, value=name)
        c = ws.cell(row=rr, column=3, value=rule)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["C"].width = 110
    return ws


def franchise_methodology(df, rc) -> str:
    """Section 8 of methodology.md."""
    counts = df["franchise_stage"].value_counts().to_dict()
    dav, fv = rc["davidson"], rc["five_points"]
    nash = df[df["cbsa_code_or_fips"] == "34980"]
    nash_txt = "not in the footprint"
    if len(nash):
        n = nash.iloc[0]
        nash_txt = (f"pro-forma ${n['proforma_deposits_2025_kusd']/1e6:,.1f}bn of a "
                    f"${n['market_deposits_2025_kusd']/1e6:,.1f}bn market = "
                    f"{n['deposit_share_2025']*100:.1f}% deposit share, "
                    f"{n['branches_2025_sod']} branches")
    stage_rows = "\n".join(
        f"| {i+1} | **{name}** | {rule} | {counts.get(name, 0)} |"
        for i, (name, rule) in enumerate(STAGE_DEFS))
    pb = df[df["franchise_stage"].eq("Pulling Back")]
    pb_txt = "; ".join(f"{r.market_name} ({r.stage_detail})" for r in pb.itertuples()) or "none"
    cons = df[df["franchise_stage"].eq("Entrenched Leader (consolidating)")]
    cons_txt = "; ".join(f"{r.market_name} (share "
                         f"{r.deposit_share_2025*100:.1f}%, {r.share_delta_5yr_bps:+.0f}bps, "
                         f"{r.net_branch_change_5yr:+d} branches)"
                         for r in cons.itertuples()) or "none"

    # ---- 8.5 threshold sensitivity ---------------------------------------- #
    _sh = pd.to_numeric(df["deposit_share_2025"], errors="coerce")
    _sens6 = int((_sh >= 0.06).sum())
    _sens8 = int((_sh >= ENTRENCHED_DEPOSIT_SHARE).sum())
    _sens10 = int((_sh >= 0.10).sum())
    _sens_move = _sens6 - _sens10
    _near = df[(_sh >= 0.06) & (_sh < 0.10)]
    _sens_near = "; ".join(f"{r.market_name} {r.deposit_share_2025*100:.1f}%"
                           for r in _near.sort_values(
                               "deposit_share_2025", ascending=False).itertuples()) or "none"

    # ---- 8.4 / 8.6b named cases -------------------------------------------- #
    def _bps_of(code):
        s = df[df["cbsa_code_or_fips"].astype(str) == code]
        return f"{s.iloc[0]['share_delta_5yr_bps']:+.1f}bps" if len(s) else "n/a"
    _col_bps, _atl_bps = _bps_of("17980"), _bps_of("12060")

    # ---- 8.6b relocation table --------------------------------------------- #
    _rl = df[(pd.to_numeric(df.get("book_moved_out_kusd", 0), errors="coerce").fillna(0) > 0)
             | (pd.to_numeric(df.get("book_moved_in_kusd", 0), errors="coerce").fillna(0) > 0)]
    _reloc_rows = "\n".join(
        f"| {r.cbsa_code_or_fips} {r.market_name} | {r.franchise_stage} | "
        f"{('$%.2fbn' % (r.book_moved_out_kusd/1e6)) if r.book_moved_out_kusd else '--'} | "
        f"{('$%.2fbn' % (r.book_moved_in_kusd/1e6)) if r.book_moved_in_kusd else '--'} | "
        f"{r.share_delta_5yr_bps:+.1f} |"
        for r in _rl.sort_values("book_moved_out_kusd", ascending=False).itertuples())
    _reloc_table = ("| Market | Stage | Book out | Book in | share_delta_5yr_bps |\n"
                    "|---|---|---|---|---|\n" + _reloc_rows) if len(_rl) else "none"

    # ---- 8.6c deposit-drop flag list --------------------------------------- #
    _dr = df[df["data_flags"].fillna("").str.contains("deposit_book_drop")]
    _drop_txt = "; ".join(
        f"{r.market_name} ({(r.proforma_deposits_2025_kusd/r.proforma_deposits_2020_kusd-1)*100:+.0f}%, "
        f"net {r.net_branch_change_5yr:+d} branches, {r.franchise_stage})"
        for r in _dr.sort_values("proforma_deposits_2020_kusd",
                                 ascending=False).itertuples()) or "none"

    return f"""
## 8. Franchise stage -- "where is newco on its journey in this market?"

Sections 1-7 describe the MARKET. This section describes the BANK'S POSITION in
that market. It is a separate dimension: a Low-growth market can be an Entrenched
Leader franchise and a High-growth market can be a two-branch New Entrant.

### 8.1 Sources

| Input | File | Content |
|---|---|---|
| Pro-forma branch/deposit history | `data/sod_history.csv` | FDIC SOD, certs **35583** (legacy Pinnacle Bank) and **873** (Synovus Bank), every survey year **1994-2025**, aggregated to cert-year-county (2,134 rows). |
| All-institution market denominator | `data/sod_market_county.csv` | FDIC SOD county totals for **every institution**, June-2020 and June-2025, for the 151 footprint counties: `market_deposits_kusd`, `market_branches`, `n_institutions` (302 county-year rows). |
| Top-5 competitor detail | `data/sod_market_detail_top.csv` | Per county-year, the five largest certs by deposits (1,488 rows; 14 county-years have fewer than five certs). Context only -- not an input to any stage rule. |
| Current office file | `data/footprint.csv` | 429 offices with `acquired_2026` provenance and `SERVTYPE` full/limited-service split. |

### 8.2 Pro-forma treatment

Certs 35583 and 873 are **summed in every year**. The January-2026 merger means the
combined company's history in a market is the **earlier** of the two predecessors'
histories: Columbus GA is not a 2026 entry for newco, it is a 30-year franchise
that newco now owns. No back-adjustment, no pro-rating -- a simple union of the
two charters' SOD footprints.

**Geography.** Counties map to the same 81 markets used above (77 CBSAs via the
2023 crosswalk + 4 non-CBSA rural counties). Two maps are used, deliberately:

* **Strict (deposit-taking footprint counties)** for every **deposit figure and
  share**. Numerator and denominator cover the same counties in the same year, so a
  share is a true within-market share -- of the counties the bank actually operates
  in, not of the full CBSA.
* **Wide (footprint counties + any other county of the same CBSA)** for
  `entry_year` **and for the bank's own branch counts**. A branch count is a
  bank-side census, not a ratio, so it must cover the whole CBSA.

**Correction applied in this run (branch counts).** The strict map was previously
used for `branches_2020` as well. Three markets held an office in 2020 in a CBSA
county they have since exited, so their 2020 count was understated by one and
`net_branch_change_5yr` overstated by +1:

| Market | Wide-only county | 2020 branches / deposits | Effect |
|---|---|---|---|
| 36740 Orlando-Kissimmee-Sanford FL | 12069 Lake FL | 1 / $16,849k | net +1 -> **0**; the market no longer qualifies as 'Expanding' |
| 45300 Tampa-St. Petersburg-Clearwater FL | 12053 Hernando FL | 1 / $81,880k | net -1 -> -2; label unchanged |
| 24860 Greenville-Anderson-Greer SC | 45007 Anderson SC | 1 / $27,251k | net 0 -> -1; label unchanged |

Orlando's 'Expanding' label rested solely on the young-market clause
(tenure < {EXPANDING_YOUNG_TENURE} AND bps > 0 AND net > 0); with the true 2020
count that clause fails. The three markets carry
`branches_2020_includes_cbsa_county_since_exited` in `data_flags`. **No 2025 SOD
branch sits in a wide-only county**, so `branches_2025_sod` is identical on both
maps and `branch_share_2025` stays numerator/denominator consistent.

**Correction applied in this run (LPO-only counties in the denominator).** The
denominator was previously summed over every footprint county, including counties
whose only presence is a loan-production or limited-service office. An LPO cannot
file a Summary of Deposits, so those counties contributed a structurally **zero
numerator against a full denominator** and diluted the share. Nine counties are
LPO-only; five constitute their whole market (those are 'Limited-Service Foothold'
and now report a **null**, not a zero, share). Four sat inside full-service markets:

| Market | LPO-only county | Share of old denominator | deposit_share_2025 old -> new |
|---|---|---|---|
| 31340 Lynchburg VA | 51680 Lynchburg city | 54.5% | 2.115% -> 4.644% |
| 40220 Roanoke VA | 51067 Franklin | 15.2% | 11.278% -> 13.299% |
| 32820 Memphis TN-MS-AR | 28033 DeSoto MS, 47167 Tipton TN | 14.9% | 6.605% -> 7.758% |
| 16740 Charlotte NC-SC | 45091 York SC | 1.0% | 0.422% -> 0.426% |

Pro-forma deposits in all nine counties are $0 in both endpoint years (verified in
`sod_history.csv`), so nothing leaves the numerator and the share stays consistent
-- it is now taken over the counties where a share is definable at all. No stage
label flips: Memphis at 7.76% still sits below the {ENTRENCHED_DEPOSIT_SHARE*100:.0f}% cutoff.

**Pre-2010 cert-873 caveat.** Before 2010 cert 873 is only the legacy Columbus Bank
& Trust charter (18-37 branches in 2-5 Georgia counties); the rest of Synovus sat
under ~30 sibling certs that consolidated into 873 in 2010 (19 -> 325 branches,
2 -> 98 counties). Pre-2010 pro-forma branch counts are therefore a **floor**, and
`entry_year` for legacy-Synovus markets is a **latest-possible** entry date -- the
true franchise is at least as old, never younger.

The bias therefore runs toward **short** tenure, and short tenure is exactly what
feeds the New Entrant clause (tenure < {NEW_ENTRANT_TENURE}) and the young-market
Expanding clause (tenure < {EXPANDING_YOUNG_TENURE}). (An earlier draft of this
note asserted the opposite -- that "nothing in the stage rules rewards an early
entry_year, so the bias is conservative" -- which is backwards.) What actually
protects the labels is the 2010 consolidation floor: the earliest spurious
`entry_year` any legacy-Synovus market can take is **2010**, giving tenure 15,
which clears both gates with room. No market is mislabelled by this, but the
reason is the 2010 floor, not the direction of the bias.

**`entry_year` can be discontinuous.** It is a `min` over survey years, so a
missing mid-year cannot reset tenure -- but `tenure_years` is elapsed-since-first-
presence, **not continuous presence**, and the pro-forma union can date a market to
a predecessor that has since left it. Three markets have interior gaps: Atlanta
12060 (absent 2000-2009), Macon 31420 (2014-2018) and Memphis 32820 (2014-2015).
Memphis is the substantive one: cert 873 ran 9 Memphis branches in 2010 falling to
6 by 2013 and then exited; cert 35583 first appears in 2016. The row reads
"entered 2010, tenure 15" when neither charter had a Memphis branch in 2014-2015
and the current franchise is Pinnacle's from 2016 (true continuous tenure 9). No
stage flips -- tenure 15 clears both the <{NEW_ENTRANT_TENURE} and
<{EXPANDING_YOUNG_TENURE} gates either way -- but read `entry_year` as
first-ever-presence, not as an unbroken franchise age.

### 8.3 Metric definitions

| Column | Definition |
|---|---|
| `entry_year` | Earliest SOD year (1994-2025) in which either cert reported >=1 branch in the market (wide county map). |
| `tenure_years` | {SOD_CURRENT_YEAR} - `entry_year`. |
| `branches_2020`, `branches_2025_sod` | Pro-forma count of **deposit-reporting SOD offices** (both certs), **wide** county map (whole CBSA). These differ from `pinnacle_offices` in section 1, which counts the Aug-2026 office file including limited-service offices, and `branches_2020` also differs from a strict-map count in the three markets tabled in 8.2. |
| `office_file_deposits_2025_kusd` vs `proforma_deposits_2025_kusd` | **Different bases; they do not tie.** The first is the office-file basis used by sections 1-7, the second the SOD basis used by every metric in this section. See section 7 for the reconciliation and the Bowling Green $0-vs-$172,876k case. Nothing in section 8 uses the office-file column. |
| `book_moved_out_kusd`, `book_moved_in_kusd` | Gross office-level book relocation detected between the endpoint surveys (section 8.6). Diagnostic only -- no stage rule reads them -- but a non-zero value means the share move behind the label is partly a booking change. |
| `branches_2020_wide_only` | Non-empty when `branches_2020` includes an office in a CBSA county the bank has since exited (8.2). |
| `net_branch_change_5yr` | `branches_2025_sod` - `branches_2020`. |
| `branch_change_pct` | `branches_2025_sod` / `branches_2020` - 1; null when `branches_2020` = 0. |
| `deposit_share_2025` | pro-forma deposits / all-institution `market_deposits_kusd`, June-2025, same counties. `deposit_share_2020` likewise. |
| `share_delta_5yr_bps` | (`deposit_share_2025` - `deposit_share_2020`) x 10,000. |
| `branch_share_2025` | `branches_2025_sod` / all-institution `market_branches` (2025). |
| `branch_effectiveness_index` (BEI) | `deposit_share_2025` / `branch_share_2025`; null when branch share is 0. **BEI > 1** = the franchise gathers more deposit share than its physical footprint would imply; **BEI < 1** = an execution or vintage gap. |
| `entered_via_2026_merger` | True when **all** current full-service branches in the market carry `acquired_2026 = True` (for the limited-service-only markets, all offices). |
| `n_institutions` | June-2025 distinct-cert count from the county file. For multi-county markets this is a **sum of per-county counts**, so a bank present in two counties of one market is counted twice; read it as competitive density ("institution-county presences"), not a distinct bank count. |

### 8.4 Stage rules (deterministic, ordered, first match wins)

| Order | Stage | Rule | Markets |
|---|---|---|---|
{stage_rows}

The table above is printed in **true evaluation order**. (An earlier draft listed
'Entrenched Leader (consolidating)' as a trailing variant *below* the
'Everything else' catch-all, in a table whose entire premise is order.)

**Borderline convention.** Rules are evaluated strictly in order and the first
match wins; no market can carry two stages. The one explicit collision is a market
that satisfies both the Pulling-Back test and the >= {ENTRENCHED_DEPOSIT_SHARE*100:.0f}%
entrenchment test. Those are labelled **'Entrenched Leader (consolidating)'**
rather than 'Pulling Back'. The {len(cons)} market(s) that hit this collision in
this run: {cons_txt}.

**Do not read 'consolidating' as a uniform story.** In most of these markets the
share slipped while a dominant franchise thinned overlapping branches. In at least
one it did not: **Columbus GA-AL (17980)** carries the largest share fall in the
set ({_col_bps}) not because two branches closed but because the Synovus corporate
book of record moved out of the market -- see 8.6. Its `stage_detail` and
`data_flags` say so on the row. The label is a rules output; the narrative behind
each one has to be read off the row, not assumed from the stage name.

**Pulling Back markets in this run:** {pb_txt}.

**Fragile labels.** Two markets sit on a knife edge and should be read with the
inputs, not the label:

* **Naples-Marco Island FL** is 'Pulling Back' on a branch change of *exactly*
  {PULLBACK_BRANCH_PCT*100:.0f}.0% (4 -> 3 branches), which passes only because the
  test is inclusive; `net_branch_change_5yr` = -1 fails the primary clause. Its
  supporting share fall rests on a deposit decline that the run's own
  `deposit_book_drop_*` flag says needs investigating. One branch, or one resolved
  booking question, moves it to 'Established - Mature'.
* **Orlando FL** was 'Expanding' before the 8.2 branch-count correction and is
  'Established - Mature' after it, on a single 2020 branch in Lake County.

### 8.5 The 8% entrenchment convention -- provenance

**The 8% cutoff is an in-house analyst convention. It is not sourced, and this
document no longer claims that it is.**

An earlier draft attributed a set of highly specific, checkable claims -- "plotted
deposit share against branch share across 250+ US counties", "an S-shaped
relationship in more than 80% of markets", "critical mass typically begins around
6-8% branch share" -- to McKinsey, *The Future of Retail Banking*, 2010, chapter
"Big Fish in Small Ponds", with **no URL, no page, no exhibit number and no
corresponding source file anywhere in this repository**. Nothing in `data/` or in
the fetch scripts supports those figures; they existed only as hardcoded prose in
`market_categorizer.py`. A parallel claim that JPMorgan targets "roughly 10%
**branch** share" in expansion markets was equally uncited, and JPMorgan's public
expansion-market statements are normally framed in **deposit** share -- so that
sentence may well have carried the very units error the text elsewhere warns
against. All of it has been removed rather than dressed up.

What survives is the reasoning, which stands on its own and does not need a
citation:

* Local deposit markets exhibit scale effects. Below some level of local presence
  a bank is a price-taker on deposits and cannot amortise local marketing,
  staffing and commercial-banking coverage; above it, share tends to compound.
  That much is uncontroversial; **where the inflection sits is market-specific and
  is not estimated here**.
* 8% is chosen as a **legible screening line**, not an estimate of that
  inflection. It answers "is this a market where we are structurally significant?"
  with a round number a reader can re-cut.
* The units caveat is real and is the reason `branch_effectiveness_index` is on
  every row: our cutoff is on **deposit** share, and BEI = deposit share / branch
  share is the bridge to any branch-share-based reference a reader wants to bring.
* Any such literature would predate mobile-first deposit gathering, digital-only
  competitors and branch-network thinning, all of which weaken the branch-share ->
  deposit-share link.

**Sensitivity to the cutoff.** Because the number is a convention, here is what
moving it does. Markets at or above the cutoff:

| Cutoff | Markets | Share of the 81-market book |
|---|---|---|
| 6% | {_sens6} | {_sens6/len(df)*100:.0f}% |
| **8% (used)** | **{_sens8}** | **{_sens8/len(df)*100:.0f}%** |
| 10% | {_sens10} | {_sens10/len(df)*100:.0f}% |

{_sens_move} markets -- {_sens_move/len(df)*100:.0f}% of the book -- change side on
a +/-2pt move. Markets within 2 points of the line, i.e. the ones a reader should
sanity-check by hand: {_sens_near}.

### 8.6 Known data wart -- main-office deposit reallocation

Deposits in the SOD are assigned to the **branch of record**, and banks periodically
move a book between offices. Cert 35583's Nashville main office (UNINUM 82012)
reported **DEPSUMBR = $0** in June-2025 after reporting $8.27bn in June-2020.

**Share levels keep SOD as-reported** -- the money did not leave the bank, and if it
moved to another office in the same county the market share is unaffected. This run
therefore *checks where the book sits* rather than assuming:

* Cert 35583, Davidson County TN (47037): **${dav['dep_2020_kusd']:,}k across
  {dav['branches_2020']} branches in 2020 -> ${dav['dep_2025_kusd']:,}k across
  {dav['branches_2025']} branches in 2025** ({dav['pct_change']*100:+.1f}%). The
  county book **{'grew, so the reallocation stayed inside Davidson County and the Nashville MSA share is NOT distorted' if dav['book_retained_in_county'] else 'FELL, so the book left the county and the Nashville share is UNDERSTATED -- read the Nashville row with that caveat'}**.
  Nashville MSA pro-forma position on this run: {nash_txt}, consistent with the
  bank's ~$21bn Nashville franchise once the offices outside Davidson County are
  included.
* Five Points branch (UNINUM {fv.get('uninum')}, cert {fv.get('cert')}, county
  {fv.get('county')}): 2025 SOD {'reports $0' if fv.get('zero_reported_2025') else 'reports a positive book'};
  the county-level cert-873 book went ${fv.get('county_dep_2020_kusd', 0):,}k ->
  ${fv.get('county_dep_2025_kusd', 0):,}k, i.e. the book
  {'stayed inside the county and the market share is unaffected' if fv.get('book_retained_in_county') else 'did not stay in the county -- treat that market share as understated'}.

### 8.6b The Columbus GA -> Atlanta corporate-book relocation

The Nashville case above is an **intra-county** move: it changes which office
reports the money, not which market. Between the 2020 and 2021 surveys the same
thing happened **across markets**, and the previous version of this document did
not detect it, name it, or flag it.

Market-level pro-forma deposits:

| Market | 2020 | 2021 | 2025 |
|---|---|---|---|
| 17980 Columbus GA-AL | $8,959,045k | $4,914,458k | $5,699,765k |
| 12060 Atlanta GA | $6,685,981k | $10,868,774k | $15,331,595k |

Office level (`data/sod_history_office.csv`), cert 873:

| UNINUM | Office | Market | 2020 | 2021 | 2025 |
|---|---|---|---|---|---|
| 556 | Synovus main office, Columbus | 17980 | $7,672,011k | $3,431,780k | $0 |
| 672980 | replacement Columbus office of record | 17980 | -- | -- | $4,221,214k |
| 591421 | OVERTON BRANCH (Cobb Cty GA) | 12060 | $633,543k | $3,753,597k | $7,888,503k |

Synovus moved its corporate book of record out of Columbus. Roughly $4.2bn of it
re-appears at a replacement Columbus office; the Overton office in metro Atlanta
picks up **+$7.25bn** over the five years, against a starting balance of $634m.
This is a booking event, not a customer migration 100 miles up the interstate.

**Consequences for two headline labels, stated plainly:**

* **Columbus is 'Entrenched Leader (consolidating)'** on
  `share_delta_5yr_bps` = {_col_bps}, the largest share fall in the book. The
  'consolidating' story -- a dominant franchise pruning overlapping branches -- is
  the wrong explanation here. Two branches closed; the share fell because the book
  moved markets.
* **Atlanta is 'Expanding'** on the >= +{EXPANDING_SHARE_BPS:.0f}bps clause with
  {_atl_bps}, while **cert 873's own Atlanta branch count FELL 45 -> 41** over the
  same period as its Atlanta deposits went $6.66bn -> $14.27bn. (Pro-forma
  `branches_2025_sod` rises to 47 only because legacy Pinnacle added 5 offices.)
  That is precisely the "market shows Expanding from a booking artifact" failure
  this section exists to prevent, and Columbus is its mirror image on the retreat
  side.

**What has changed in this run.** An office-level detector now runs over the two
endpoint surveys ({SOD_BASE_YEAR} vs {SOD_CURRENT_YEAR}). An office is reported as
a material book movement only when it is **both**

* material in dollars -- a move of at least
  **${RELOC_OFFICE_MIN_KUSD/1e6:,.1f}bn**, or at least
  **{RELOC_OFFICE_MIN_FRAC*100:.0f}% of its market's own {SOD_BASE_YEAR} book**
  subject to a ${RELOC_OFFICE_FLOOR_KUSD/1e3:,.0f}m floor; **and**
* anomalous in its own terms -- a **source** kept
  <={RELOC_SOURCE_RETAINED*100:.0f}% of its own {SOD_BASE_YEAR} balance, a
  **recipient** ended at >={RELOC_RECIPIENT_MULT:.0f}x its own {SOD_BASE_YEAR}
  balance (an office that did not exist in {SOD_BASE_YEAR} qualifies only on the
  absolute ${RELOC_OFFICE_MIN_KUSD/1e6:,.1f}bn test).

The second condition is what separates a booking move from a large branch simply
growing: size alone is not evidence. Every affected market -- **source and
recipient alike** --
carries `book_relocation_out:` / `book_relocation_in:` in `data_flags` and a
`CAUTION:` clause appended to `stage_detail`, next to the label the movement drove.
Markets flagged in this run:

{_reloc_table}

The labels themselves are left as the rules produce them. The stage rules are a
deterministic screen, not a judgement (8.7); the fix for an artifact-driven label
is to make the artifact impossible to miss on the row, which is now the case.

### 8.6c The blanket deposit-drop flag

Any market whose pro-forma deposits fell more than
{abs(REALLOC_DROP_FLAG)*100:.0f}% between the endpoint surveys is flagged
`deposit_book_drop_*:check_main_office_reallocation`.

**This test used to be switched off exactly where it was needed.** It previously
fired only when `net_branch_change_5yr > -2`, which exempted the two largest
pro-forma deposit collapses in the book -- **Columbus GA-AL** (deposits -36.4%,
net -2) and **Miami-Fort Lauderdale-West Palm Beach FL** (-29.4%, net -3) -- both
markets whose stage label is driven by the very deposit movement the flag exists
to question. The guard also inverted the logic: a bank that closes branches *and*
loses a third of its book is the least plausible case to exempt from a
booking-artifact check, not the most. The branch condition has been removed.
Markets flagged now: {_drop_txt}.

### 8.7 Limitations of the franchise-stage dimension

* **Annual June snapshots.** The SOD is a single as-of-June-30 observation each
  year. Branches opened and closed inside a survey year are invisible, and a
  quarter-end balance-sheet position is not an average balance.
* **Deposits are assigned to the branch of record.** Commercial, brokered, trust
  and digital deposits are booked where the relationship sits, not where the
  customer is; large-corporate books concentrate at HQ. Section 8.6 names the
  Nashville (intra-county) and Columbus GA -> Atlanta (cross-market) cases, and an
  office-level detector now flags both sides of any move above the 8.6b threshold
  -- but the effect is general and continuous, and only the large discrete moves
  are detectable this way. It inflates every HQ market's share while deflating its
  neighbours'.
* **The 8% entrenchment line is a convention, not an estimate**, and it is not
  sourced to any literature (8.5). {_sens_move} of the {len(df)} markets change
  side on a +/-2pt move.
* **Two kinds of zero share.** The four Limited-Service Foothold markets now
  report `deposit_share_2025` as **null**, not 0.0: an LPO cannot file an SOD, so
  the share is undefined and averaging or ranking the column no longer silently
  includes them. Decatur AL is different and is left as a **measured** 0.0% -- it
  has a full-service branch, opened 27-Jan-2026, and genuinely held no deposits at
  the June-2025 survey date. Its row carries
  `no_sod_history:entry_year_unavailable` and reads "entered n/a".
* **Same-county denominators.** Shares are computed over the counties the bank
  operates in, not the full CBSA. For a market where the bank sits in one county
  of a five-county metro, the share answers "share where we are", which is the
  right operating question but is NOT comparable to a published full-MSA share.
* **`entry_year` is a SOD-presence date**, not a charter or first-customer date,
  and for pre-2010 legacy-Synovus markets it is a latest-possible bound (8.2).
* **Stage is a rules-based label, not a judgement.** The thresholds are round
  numbers chosen for legibility; a market one branch either side of a cutoff can
  flip stage. `stage_detail` exposes every input on the row so a reader can
  overrule the label without re-running anything.
* **No competitor dynamics.** The rules use only the bank's own trajectory and its
  share of the market total; a share gain taken from a failing competitor and one
  taken from a healthy one look identical here.
"""


if __name__ == "__main__":
    res = main()
    df = res["df"]
    out = {
        "rows": len(df),
        "tier_counts": df["growth_tier"].value_counts().to_dict(),
        "size_counts": df["size_tier"].value_counts().to_dict(),
    }
    print(json.dumps(out, indent=1))
    for code in ["34980", "12060", "13820", "33100"]:
        r = df[df.cbsa_code_or_fips == code]
        if len(r):
            r = r.iloc[0]
            print(f"{code} {r.market_name[:40]:42s} {r.size_tier:16s} "
                  f"{r.growth_tier:7s} {r.composite_z:+.3f} pop={r.pop_2025:,}")
    for _, r in df[df.size_tier == "Rural"].iterrows():
        print(f"RURAL {r.cbsa_code_or_fips} {r.market_name[:34]:36s} "
              f"{r.growth_tier:7s} {r.composite_z:+.3f}")
